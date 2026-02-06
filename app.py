from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import ProgrammingError
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
from functools import wraps
import os
import hashlib
import random
import string

app = Flask(__name__)
app.config['SECRET_KEY'] = 'sua-chave-secreta-super-segura-123'

# ==================== MIDDLEWARE DE VERIFICAÇÃO DE LICENÇA ====================

# Define o caminho absoluto para o banco de dados
basedir = os.path.abspath(os.path.dirname(__file__))
db_path = os.path.join(basedir, 'database')
# Criar pasta database se não existir
os.makedirs(db_path, exist_ok=True)
# Render/Prod: usar DATABASE_URL (PostgreSQL). Dev local: SQLite.
# Railway/Render podem usar nomes diferentes de env.
database_url = (
    os.environ.get('DATABASE_URL')
    or os.environ.get('POSTGRES_URL')
    or os.environ.get('POSTGRES_PRISMA_URL')
    or os.environ.get('POSTGRES_URL_NON_POOLING')
)
if database_url:
    database_url = database_url.strip().strip('"').strip("'")
if database_url:
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    if database_url.startswith('postgresql+psycopg2://'):
        database_url = database_url.replace('postgresql+psycopg2://', 'postgresql+psycopg://', 1)
    elif database_url.startswith('postgresql://'):
        database_url = database_url.replace('postgresql://', 'postgresql+psycopg://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(db_path, 'caixa.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ==================== DB INIT (AUTO) ====================

def _ensure_db_ready():
    with app.app_context():
        db.create_all()
        # Correção automática: Adicionar coluna acesso_relatorios se não existir
        try:
            from sqlalchemy import text
            with db.engine.connect() as conn:
                conn.execute(text('ALTER TABLE usuario ADD COLUMN acesso_relatorios BOOLEAN DEFAULT 0'))
                conn.commit()
        except Exception:
            pass
        # Criar registros padrão (inclui ADMIN MASTER e ADMIN)
        try:
            init_db()
        except Exception:
            pass

# Executar no startup para evitar erro de tabela inexistente no primeiro request
try:
    _ensure_db_ready()
except Exception:
    pass

_db_initialized = False

@app.before_request
def ensure_db_initialized():
    global _db_initialized
    if _db_initialized:
        return
    try:
        _ensure_db_ready()
        _db_initialized = True
    except Exception:
        pass

# ==================== CONTEXT PROCESSORS ====================

@app.context_processor
def inject_user():
    """Injeta informações do usuário em todos os templates"""
    if 'user_id' in session:
        usuario = db.session.get(Usuario, session['user_id'])
        return dict(usuario_logado=usuario)
    return dict(usuario_logado=None)

@app.context_processor
def inject_datetime():
    from datetime import datetime
    return dict(datetime=datetime)

# No contexto do processador, adicione a função now()
@app.context_processor
def utility_processor():
    """Injeta funções utilitárias em todos os templates"""
    def calcular_totais_caixa_template(caixa):
        return calcular_totais_caixa(caixa)
    def now():
        return datetime.utcnow()
    def format_currency(value):
        return formatar_moeda(value)
    return dict(
        calcular_totais_caixa=calcular_totais_caixa_template, 
        now=now,
        format_currency=format_currency
    )


# Adicione esta função no seu app.py
def formatar_moeda(valor):
    """Formata valor para moeda brasileira R$ 1.234,56"""
    if valor is None:
        return "R$ 0,00"
    try:
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return f"R$ {valor}"

def parse_moeda(valor, default=0.0):
    """Converte string moeda BR (ex: 1.234,56) para float"""
    if valor is None:
        return default
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip()
    if s == '':
        return default
    s = s.replace('R$', '').replace(' ', '')
    s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except Exception:
        return default



# ==================== MODELS ====================

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    senha = db.Column(db.String(200), nullable=False)
    perfil = db.Column(db.String(20), default='OPERADOR')
    acesso_dashboard = db.Column(db.Boolean, default=True)
    acesso_configuracoes = db.Column(db.Boolean, default=False)
    acesso_relatorios = db.Column(db.Boolean, default=False)  # NOVO CAMPO
    ativo = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

class Caixa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.Date, nullable=False)
    turno = db.Column(db.String(20), nullable=False)
    operador_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    saldo_inicial = db.Column(db.Float, default=0)
    saldo_final = db.Column(db.Float, default=0)
    status = db.Column(db.String(20), default='ABERTO')
    hora_abertura = db.Column(db.DateTime, default=datetime.utcnow)
    hora_fechamento = db.Column(db.DateTime)
    operador = db.relationship('Usuario', backref='caixas')

class FormaPagamento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(50), nullable=False, unique=True)
    ativo = db.Column(db.Boolean, default=True)

class BandeiraCartao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(50), nullable=False, unique=True)
    ativo = db.Column(db.Boolean, default=True)

class CategoriaDespesa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    tipo = db.Column(db.String(20), nullable=False)  # FIXA, VARIAVEL, SAIDA
    ativo = db.Column(db.Boolean, default=True)

class Motoboy(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    taxa_padrao = db.Column(db.Float, default=5.00)
    ativo = db.Column(db.Boolean, default=True)

class Venda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    caixa_id = db.Column(db.Integer, db.ForeignKey('caixa.id'))
    tipo = db.Column(db.String(20), nullable=False)  # MESA, BALCAO
    numero = db.Column(db.Integer)
    total = db.Column(db.Float, nullable=False)
    emitiu_nota = db.Column(db.Boolean, default=False)
    observacao = db.Column(db.Text)
    data_hora = db.Column(db.DateTime, default=datetime.utcnow)
    caixa = db.relationship('Caixa', backref='vendas')

class PagamentoVenda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    venda_id = db.Column(db.Integer, db.ForeignKey('venda.id'))
    forma_pagamento_id = db.Column(db.Integer, db.ForeignKey('forma_pagamento.id'))
    bandeira_id = db.Column(db.Integer, db.ForeignKey('bandeira_cartao.id'))
    valor = db.Column(db.Float, nullable=False)
    observacao = db.Column(db.String(200))
    venda = db.relationship('Venda', backref='pagamentos')
    forma_pagamento = db.relationship('FormaPagamento')
    bandeira = db.relationship('BandeiraCartao')

class Delivery(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    caixa_id = db.Column(db.Integer, db.ForeignKey('caixa.id'))
    cliente = db.Column(db.String(100), nullable=False)
    total = db.Column(db.Float, nullable=False)
    taxa_entrega = db.Column(db.Float, default=0)
    motoboy_id = db.Column(db.Integer, db.ForeignKey('motoboy.id'))
    emitiu_nota = db.Column(db.Boolean, default=False)
    observacao = db.Column(db.Text)
    data_hora = db.Column(db.DateTime, default=datetime.utcnow)
    caixa = db.relationship('Caixa', backref='deliveries')
    motoboy = db.relationship('Motoboy')

class PagamentoDelivery(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    delivery_id = db.Column(db.Integer, db.ForeignKey('delivery.id'))
    forma_pagamento_id = db.Column(db.Integer, db.ForeignKey('forma_pagamento.id'))
    bandeira_id = db.Column(db.Integer, db.ForeignKey('bandeira_cartao.id'))
    valor = db.Column(db.Float, nullable=False)
    observacao = db.Column(db.String(200))
    delivery = db.relationship('Delivery', backref='pagamentos')
    forma_pagamento = db.relationship('FormaPagamento')
    bandeira = db.relationship('BandeiraCartao')

class Despesa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    caixa_id = db.Column(db.Integer, db.ForeignKey('caixa.id'))
    tipo = db.Column(db.String(20), nullable=False)  # FIXA, VARIAVEL, SAIDA
    categoria_id = db.Column(db.Integer, db.ForeignKey('categoria_despesa.id'))
    descricao = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Float, nullable=False)
    forma_pagamento_id = db.Column(db.Integer, db.ForeignKey('forma_pagamento.id'))
    data_vencimento = db.Column(db.Date)
    status = db.Column(db.String(20), default='PAGO')
    observacao = db.Column(db.Text)
    data_hora = db.Column(db.DateTime, default=datetime.utcnow)
    caixa = db.relationship('Caixa', backref='despesas')
    categoria = db.relationship('CategoriaDespesa')
    forma_pagamento = db.relationship('FormaPagamento')

class Sangria(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    caixa_id = db.Column(db.Integer, db.ForeignKey('caixa.id'))
    valor = db.Column(db.Float, nullable=False)
    motivo = db.Column(db.String(100), nullable=False)
    observacao = db.Column(db.String(200))
    data_hora = db.Column(db.DateTime, default=datetime.utcnow)
    caixa = db.relationship('Caixa', backref='sangrias')

# ==================== MODELO SUPRIMENTO (v3.0) ====================
class Suprimento(db.Model):
    """Suprimentos - Entradas de dinheiro no caixa"""
    __tablename__ = 'suprimento'
    id = db.Column(db.Integer, primary_key=True)
    caixa_id = db.Column(db.Integer, db.ForeignKey('caixa.id'))
    valor = db.Column(db.Float, nullable=False)
    motivo = db.Column(db.String(100), nullable=False)
    observacao = db.Column(db.String(200))
    data_hora = db.Column(db.DateTime, default=datetime.utcnow)
    caixa = db.relationship('Caixa', backref='suprimentos')

class Produto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), unique=True)
    nome = db.Column(db.String(200), nullable=False)
    categoria = db.Column(db.String(50))
    custo = db.Column(db.Float, default=0)
    preco_venda = db.Column(db.Float, default=0)
    quantidade = db.Column(db.Integer, default=0)
    estoque_minimo = db.Column(db.Integer, default=0)
    estoque_maximo = db.Column(db.Integer, default=0)
    unidade = db.Column(db.String(10), default='UN')
    ativo = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

class MovimentacaoEstoque(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    produto_id = db.Column(db.Integer, db.ForeignKey('produto.id'))
    tipo = db.Column(db.String(20), nullable=False)  # ENTRADA, SAIDA, AJUSTE
    quantidade = db.Column(db.Integer, nullable=False)
    valor_unitario = db.Column(db.Float, default=0)
    valor_total = db.Column(db.Float, default=0)
    motivo = db.Column(db.String(100))
    observacao = db.Column(db.Text)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    data_hora = db.Column(db.DateTime, default=datetime.utcnow)
    produto = db.relationship('Produto', backref='movimentacoes')
    usuario = db.relationship('Usuario')

# ==================== MODELOS DE LICENÇA ====================

class Licenca(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(100), nullable=False, unique=True)
    chave_ativacao = db.Column(db.String(50), nullable=False, unique=True)
    data_ativacao = db.Column(db.DateTime, default=datetime.utcnow)
    data_expiracao = db.Column(db.DateTime)
    status = db.Column(db.String(20), default='ATIVA')  # ATIVA, EXPIRADA, BLOQUEADA
    max_dispositivos = db.Column(db.Integer, default=2)
    ativo = db.Column(db.Boolean, default=True)

class Dispositivo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    licenca_id = db.Column(db.Integer, db.ForeignKey('licenca.id'))
    nome = db.Column(db.String(100))
    endereco_ip = db.Column(db.String(50))
    mac_address = db.Column(db.String(50))
    user_agent = db.Column(db.String(200))
    data_registro = db.Column(db.DateTime, default=datetime.utcnow)
    ultimo_acesso = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='ATIVO')  # ATIVO, BLOQUEADO
    dispositivo_id = db.Column(db.String(100), unique=True)  # Hash único do dispositivo
    licenca = db.relationship('Licenca', backref='dispositivos')

class Backup(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome_arquivo = db.Column(db.String(200))
    tamanho = db.Column(db.Integer)
    data_backup = db.Column(db.DateTime, default=datetime.utcnow)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    observacao = db.Column(db.Text)
    usuario = db.relationship('Usuario')

# ==================== DECORATORS ====================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def dashboard_required(f):
    """Requer permissão para acessar dashboard"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        usuario = db.session.get(Usuario, session['user_id'])
        if not usuario or (usuario.perfil != 'MASTER' and not usuario.acesso_dashboard):
            flash('Acesso negado. Você não tem permissão para acessar o dashboard.', 'danger')
            return redirect(url_for('vendas'))
        return f(*args, **kwargs)
    return decorated_function

def estoque_required(f):
    """Requer permissão para acessar estoque (admin ou perfil específico)"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        usuario = db.session.get(Usuario, session['user_id'])
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        usuario = db.session.get(Usuario, session['user_id'])
        if not usuario or usuario.perfil not in ['ADMIN', 'MASTER']:
            flash('Acesso negado. Voce nao tem permissao para acessar esta area.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


def admin_master_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        usuario = db.session.get(Usuario, session['user_id'])
        if not usuario or usuario.perfil != 'MASTER':
            flash('Acesso negado. Apenas o ADMIN MASTER pode acessar esta area.', 'danger')
            return redirect(url_for('vendas'))
        return f(*args, **kwargs)
    return decorated_function


# ==================== ROUTES - AUTH ====================

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('vendas'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        nome = request.form.get('operador')
        senha = request.form.get('senha')
        acao = request.form.get('acao', 'novo')
        
        usuario = Usuario.query.filter_by(nome=nome, ativo=True).first()
        
        if usuario and check_password_hash(usuario.senha, senha):
            if usuario.perfil == 'MASTER':
                usuario.acesso_dashboard = True
                usuario.acesso_configuracoes = True
                usuario.acesso_relatorios = True
                db.session.commit()
            
            # ACESSAR CAIXA EXISTENTE
            if acao == 'acessar':
                caixa_id = int(request.form.get('caixa_id'))
                caixa = db.session.get(Caixa, caixa_id)
                
                if not caixa:
                    flash('Caixa não encontrado!', 'danger')
                    return redirect(url_for('login'))
                
                # Restaurar sessão com TODAS as informações
                session['user_id'] = usuario.id
                session['user_nome'] = usuario.nome
                session['caixa_id'] = caixa.id
                session['turno'] = caixa.turno
                session['data'] = caixa.data.strftime('%Y-%m-%d')
                session['saldo_inicial'] = caixa.saldo_inicial
                session['hora_abertura'] = caixa.hora_abertura.strftime('%H:%M:%S')
                
                # Calcular totais atuais
                totais = calcular_totais_caixa(caixa)
                session['total_vendas_atual'] = totais['total_vendas']
                session['total_despesas_atual'] = totais['despesas']
                session['saldo_atual'] = totais['saldo_atual']
                
                flash(f'✅ Bem-vindo de volta, {usuario.nome}! Continuando caixa #{caixa.id} - {caixa.turno}', 'success')
                return redirect(url_for('vendas'))
            
            # ABRIR NOVO CAIXA
            else:
                data = request.form.get('data')
                turno = request.form.get('turno')
                saldo_inicial = parse_moeda(request.form.get('saldo_inicial', 100))
                
                data_obj = datetime.strptime(data, '%Y-%m-%d').date()
                
                caixa_aberto = Caixa.query.filter_by(data=data_obj, turno=turno, status='ABERTO').first()
                caixa_fechado = Caixa.query.filter_by(data=data_obj, turno=turno, status='FECHADO').first()
                if usuario.perfil in ['ADMIN', 'MASTER']:
                    if caixa_aberto:
                        caixa_para_usar = caixa_aberto
                        flash(f'Admin assumindo caixa ABERTO - {turno}', 'info')
                    elif caixa_fechado:
                        caixa_fechado.status = 'ABERTO'
                        caixa_fechado.hora_fechamento = None
                        db.session.commit()
                        caixa_para_usar = caixa_fechado
                        flash(f'Caixa FECHADO reaberto pelo Admin', 'warning')
                    else:
                        caixa_para_usar = Caixa(
                            data=data_obj, turno=turno, operador_id=usuario.id,
                            saldo_inicial=saldo_inicial, status='ABERTO'
                        )
                        db.session.add(caixa_para_usar)
                        db.session.commit()
                        flash('✅ Novo caixa aberto com sucesso!', 'success')
                    
                    session['user_id'] = usuario.id
                    session['user_nome'] = usuario.nome
                    session['caixa_id'] = caixa_para_usar.id
                    session['turno'] = turno
                    session['data'] = data
                    session['saldo_inicial'] = caixa_para_usar.saldo_inicial
                    session['hora_abertura'] = caixa_para_usar.hora_abertura.strftime('%H:%M:%S')
                    return redirect(url_for('vendas'))
                else:
                    if caixa_aberto:
                        # Operador pode ENTRAR no caixa aberto
                        session['user_id'] = usuario.id
                        session['user_nome'] = usuario.nome
                        session['caixa_id'] = caixa_aberto.id
                        session['turno'] = caixa_aberto.turno
                        session['data'] = caixa_aberto.data.strftime('%Y-%m-%d')
                        session['saldo_inicial'] = caixa_aberto.saldo_inicial
                        session['hora_abertura'] = caixa_aberto.hora_abertura.strftime('%H:%M:%S')

                        flash(f'✅ Bem-vindo, {usuario.nome}! Caixa do turno {turno} acessado.', 'success')
                        return redirect(url_for('vendas'))

                    if caixa_fechado:
                        # Permitir operador acessar caixa fechado para visualização
                        session['user_id'] = usuario.id
                        session['user_nome'] = usuario.nome
                        session['caixa_id'] = caixa_fechado.id
                        session['turno'] = caixa_fechado.turno
                        session['data'] = caixa_fechado.data.strftime('%Y-%m-%d')
                        session['saldo_inicial'] = caixa_fechado.saldo_inicial
                        session['hora_abertura'] = caixa_fechado.hora_abertura.strftime('%H:%M:%S')
                        
                        flash(f'⚠️ Acessando caixa FECHADO de {turno} (Modo Visualização).', 'warning')
                        return redirect(url_for('vendas'))
                    
                    novo_caixa = Caixa(
                        data=data_obj, turno=turno, operador_id=usuario.id,
                        saldo_inicial=saldo_inicial, status='ABERTO'
                    )
                    db.session.add(novo_caixa)
                    db.session.commit()
                    
                    session['user_id'] = usuario.id
                    session['user_nome'] = usuario.nome
                    session['caixa_id'] = novo_caixa.id
                    session['turno'] = turno
                    session['data'] = data
                    session['saldo_inicial'] = novo_caixa.saldo_inicial
                    session['hora_abertura'] = novo_caixa.hora_abertura.strftime('%H:%M:%S')
                    flash('✅ Caixa aberto com sucesso!', 'success')
                    return redirect(url_for('vendas'))
        else:
            flash('❌ Credenciais inválidas!', 'danger')
    
    try:
        usuarios = Usuario.query.filter_by(ativo=True).order_by(Usuario.nome).all()
    except ProgrammingError:
        db.session.rollback()
        _ensure_db_ready()
        usuarios = Usuario.query.filter_by(ativo=True).order_by(Usuario.nome).all()
    hoje = datetime.now().date()
    caixas_abertos_hoje = Caixa.query.filter_by(data=hoje, status='ABERTO').order_by(Caixa.turno).all()
    
    return render_template('login.html', usuarios=usuarios, caixas_abertos=caixas_abertos_hoje)

@app.route('/logout')
def logout():
    session.clear()
    flash('Logout realizado com sucesso!', 'info')
    return redirect(url_for('login'))

# ==================== ROUTES - LICENCIAMENTO ====================

def _gerar_chave_ativacao():
    return '-'.join(''.join(random.choices(string.ascii_uppercase + string.digits, k=4)) for _ in range(4))

def _device_fingerprint():
    ip = request.remote_addr or '0.0.0.0'
    ua = request.user_agent.string or 'unknown'
    raw = f"{ip}|{ua}"
    return hashlib.md5(raw.encode('utf-8')).hexdigest(), ip, ua

@app.route('/ativacao', methods=['GET', 'POST'])
def ativacao():
    if request.method == 'POST':
        email = request.form.get('email')
        chave = request.form.get('chave')
        if not email or not chave:
            flash('Informe e-mail e chave para ativar.', 'warning')
            return redirect(url_for('ativacao'))

        licenca = Licenca.query.filter_by(email=email, chave_ativacao=chave).first()
        if not licenca:
            licenca = Licenca(
                email=email,
                chave_ativacao=chave,
                data_ativacao=datetime.utcnow(),
                data_expiracao=datetime.utcnow() + timedelta(days=365),
                status='ATIVA',
                ativo=True
            )
            db.session.add(licenca)
            db.session.commit()

        flash('Sistema ativado com sucesso!', 'success')
        return redirect(url_for('login'))

    return render_template('ativacao.html')

@app.route('/licenciamento')
@login_required
@admin_master_required
def licenciamento():
    licenca = Licenca.query.first()
    backups = Backup.query.order_by(Backup.data_backup.desc()).all()
    dispositivos = licenca.dispositivos if licenca else []

    dispositivo_id, ip, ua = _device_fingerprint()
    if licenca:
        dispositivo = Dispositivo.query.filter_by(dispositivo_id=dispositivo_id, licenca_id=licenca.id).first()
        if not dispositivo:
            dispositivo = Dispositivo(
                licenca_id=licenca.id,
                nome=request.user_agent.browser or 'Dispositivo',
                endereco_ip=ip,
                user_agent=ua,
                dispositivo_id=dispositivo_id,
                status='ATIVO'
            )
            db.session.add(dispositivo)
        dispositivo.ultimo_acesso = datetime.utcnow()
        db.session.commit()

    return render_template(
        'licenciamento.html',
        licenca=licenca,
        dispositivos=dispositivos,
        backups=backups,
        dispositivo_id=dispositivo_id
    )

@app.route('/licenca/registrar-dispositivo', methods=['POST'])
@login_required
@admin_master_required
def registrar_dispositivo():
    licenca = Licenca.query.first()
    if not licenca or not licenca.ativo:
        return jsonify({'status': 'no_license', 'error': 'Licença não ativa'}), 400

    dispositivo_id, ip, ua = _device_fingerprint()
    dispositivo = Dispositivo.query.filter_by(dispositivo_id=dispositivo_id, licenca_id=licenca.id).first()

    if not dispositivo:
        if len(licenca.dispositivos) >= licenca.max_dispositivos:
            return jsonify({'status': 'limit_reached', 'error': 'Limite de dispositivos atingido'}), 400
        dispositivo = Dispositivo(
            licenca_id=licenca.id,
            nome=request.user_agent.browser or 'Dispositivo',
            endereco_ip=ip,
            user_agent=ua,
            dispositivo_id=dispositivo_id,
            status='ATIVO'
        )
        db.session.add(dispositivo)

    dispositivo.ultimo_acesso = datetime.utcnow()
    db.session.commit()
    return jsonify({'status': 'ok'})

@app.route('/licenca/gerar-nova-chave', methods=['POST'])
@login_required
@admin_master_required
def gerar_nova_chave():
    licenca = Licenca.query.first()
    if not licenca:
        flash('Nenhuma licença encontrada.', 'warning')
        return redirect(url_for('licenciamento'))
    licenca.chave_ativacao = _gerar_chave_ativacao()
    licenca.status = 'ATIVA'
    licenca.ativo = True
    db.session.commit()
    flash('Nova chave gerada com sucesso!', 'success')
    return redirect(url_for('licenciamento'))

@app.route('/licenca/bloquear-todos-dispositivos', methods=['POST'])
@login_required
@admin_master_required
def bloquear_todos_dispositivos():
    licenca = Licenca.query.first()
    if licenca:
        Dispositivo.query.filter_by(licenca_id=licenca.id).update({'status': 'BLOQUEADO'})
        db.session.commit()
    flash('Todos os dispositivos foram bloqueados.', 'warning')
    return redirect(url_for('licenciamento'))

@app.route('/licenca/bloquear-dispositivo/<int:dispositivo_id>', methods=['POST'])
@login_required
@admin_master_required
def bloquear_dispositivo(dispositivo_id):
    dispositivo = db.session.get(Dispositivo, dispositivo_id)
    if dispositivo:
        dispositivo.status = 'BLOQUEADO'
        db.session.commit()
        flash('Dispositivo bloqueado.', 'warning')
    return redirect(url_for('licenciamento'))

@app.route('/licenca/desbloquear-dispositivo/<int:dispositivo_id>', methods=['POST'])
@login_required
@admin_master_required
def desbloquear_dispositivo(dispositivo_id):
    dispositivo = db.session.get(Dispositivo, dispositivo_id)
    if dispositivo:
        dispositivo.status = 'ATIVO'
        db.session.commit()
        flash('Dispositivo desbloqueado.', 'success')
    return redirect(url_for('licenciamento'))

@app.route('/licenca/excluir-dispositivo/<int:dispositivo_id>', methods=['POST'])
@login_required
@admin_master_required
def excluir_dispositivo(dispositivo_id):
    dispositivo = db.session.get(Dispositivo, dispositivo_id)
    if dispositivo:
        db.session.delete(dispositivo)
        db.session.commit()
        flash('Dispositivo excluído.', 'success')
    return redirect(url_for('licenciamento'))

@app.route('/licenca/upload-backup', methods=['POST'])
@login_required
@admin_master_required
def upload_backup():
    arquivo = request.files.get('backup')
    observacao = request.form.get('observacao', '')
    if not arquivo or arquivo.filename == '':
        flash('Selecione um arquivo de backup.', 'warning')
        return redirect(url_for('licenciamento'))
    if not arquivo.filename.lower().endswith('.db'):
        flash('Apenas arquivos .db são permitidos.', 'danger')
        return redirect(url_for('licenciamento'))

    pasta = os.path.join(basedir, 'backups')
    os.makedirs(pasta, exist_ok=True)
    nome_arquivo = secure_filename(arquivo.filename)
    caminho = os.path.join(pasta, nome_arquivo)
    arquivo.save(caminho)

    backup = Backup(
        nome_arquivo=nome_arquivo,
        tamanho=os.path.getsize(caminho),
        usuario_id=session.get('user_id'),
        observacao=observacao
    )
    db.session.add(backup)
    db.session.commit()
    flash('Backup enviado com sucesso.', 'success')
    return redirect(url_for('licenciamento'))

@app.route('/licenca/download-backup/<int:backup_id>')
@login_required
@admin_master_required
def download_backup(backup_id):
    backup = db.session.get(Backup, backup_id)
    if not backup:
        flash('Backup não encontrado.', 'warning')
        return redirect(url_for('licenciamento'))
    pasta = os.path.join(basedir, 'backups')
    return send_from_directory(pasta, backup.nome_arquivo, as_attachment=True)

@app.route('/licenca/excluir-backup/<int:backup_id>', methods=['POST'])
@login_required
@admin_master_required
def excluir_backup(backup_id):
    backup = db.session.get(Backup, backup_id)
    if backup:
        pasta = os.path.join(basedir, 'backups')
        caminho = os.path.join(pasta, backup.nome_arquivo)
        if os.path.exists(caminho):
            os.remove(caminho)
        db.session.delete(backup)
        db.session.commit()
        flash('Backup excluído.', 'success')
    return redirect(url_for('licenciamento'))

# ==================== ROUTES - VENDAS ====================

@app.route('/vendas')
@login_required
def vendas():
    caixa = db.session.get(Caixa, session['caixa_id'])
    if not caixa:
        session.clear()
        flash('Caixa não encontrado. Por favor, faça login novamente.', 'warning')
        return redirect(url_for('login'))

    vendas = Venda.query.filter_by(caixa_id=session['caixa_id']).order_by(Venda.data_hora.desc()).all()
    formas_pagamento = FormaPagamento.query.filter_by(ativo=True).all()
    bandeiras = BandeiraCartao.query.filter_by(ativo=True).all()
    
    # Calcular totais
    totais = calcular_totais_caixa(caixa)
    
    return render_template('vendas.html', 
                         vendas=vendas, 
                         formas_pagamento=formas_pagamento,
                         bandeiras=bandeiras,
                         totais=totais,
                         caixa=caixa)

@app.route('/vendas/nova', methods=['POST'])
@login_required
def nova_venda():
    try:
        tipo = request.form.get('tipo')
        numero = int(request.form.get('numero', 1))
        total = parse_moeda(request.form.get('total'))
        emitiu_nota = request.form.get('emitiu_nota') == 'on'
        observacao = request.form.get('observacao', '')
        
        # Criar venda
        venda = Venda(
            caixa_id=session['caixa_id'],
            tipo=tipo,
            numero=numero,
            total=total,
            emitiu_nota=emitiu_nota,
            observacao=observacao
        )
        db.session.add(venda)
        db.session.flush()
        
        # Adicionar pagamentos
        formas = request.form.getlist('forma_pagamento[]')
        valores = request.form.getlist('valor_pagamento[]')
        bandeiras_ids = request.form.getlist('bandeira[]')
        obs_pagamentos = request.form.getlist('obs_pagamento[]')
        
        total_pago = 0
        for i, forma_id in enumerate(formas):
            if forma_id and valores[i]:
                valor = parse_moeda(valores[i])
                if valor > 0:
                    pagamento = PagamentoVenda(
                        venda_id=venda.id,
                        forma_pagamento_id=int(forma_id),
                        valor=valor,
                        bandeira_id=int(bandeiras_ids[i]) if bandeiras_ids[i] and bandeiras_ids[i] != '' else None,
                        observacao=obs_pagamentos[i] if i < len(obs_pagamentos) else ''
                    )
                    db.session.add(pagamento)
                    total_pago += valor
        
        if abs(total_pago - total) > 0.01:
            db.session.rollback()
            flash('O total dos pagamentos não corresponde ao valor da venda!', 'danger')
            return redirect(url_for('vendas'))
        
        db.session.commit()
        flash('Venda registrada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar venda: {str(e)}', 'danger')
    
    return redirect(url_for('vendas'))

# ==================== ROUTES - DELIVERY ====================

@app.route('/delivery')
@login_required
def delivery():
    caixa = db.session.get(Caixa, session['caixa_id'])
    if not caixa:
        session.clear()
        flash('Caixa não encontrado. Por favor, faça login novamente.', 'warning')
        return redirect(url_for('login'))

    deliveries = Delivery.query.filter_by(caixa_id=session['caixa_id']).order_by(Delivery.data_hora.desc()).all()
    formas_pagamento = FormaPagamento.query.filter_by(ativo=True).all()
    motoboys = Motoboy.query.filter_by(ativo=True).all()
    
    totais = calcular_totais_caixa(caixa)
    totais_delivery = calcular_totais_delivery(caixa)
    
    bandeiras = BandeiraCartao.query.filter_by(ativo=True).all()
    return render_template('delivery.html', bandeiras=bandeiras,
                         deliveries=deliveries,
                         formas_pagamento=formas_pagamento,
                         motoboys=motoboys,
                         totais=totais,
                         totais_delivery=totais_delivery,
                         caixa=caixa)

@app.route('/delivery/novo', methods=['POST'])
@login_required
def novo_delivery():
    try:
        cliente = request.form.get('cliente')
        total = parse_moeda(request.form.get('total'))
        taxa_entrega = parse_moeda(request.form.get('taxa_entrega', 0))
        motoboy_id = request.form.get('motoboy_id')
        emitiu_nota = request.form.get('emitiu_nota') == 'on'
        observacao = request.form.get('observacao', '')
        
        delivery = Delivery(
            caixa_id=session['caixa_id'],
            cliente=cliente,
            total=total,
            taxa_entrega=taxa_entrega,
            motoboy_id=int(motoboy_id) if motoboy_id else None,
            emitiu_nota=emitiu_nota,
            observacao=observacao
        )
        db.session.add(delivery)
        db.session.flush()
        
        # Adicionar pagamentos
        formas = request.form.getlist('forma_pagamento_delivery[]')
        valores = request.form.getlist('valor_pagamento_delivery[]')
        obs_pagamentos = request.form.getlist('obs_pagamento_delivery[]')
        
        total_com_taxa = total + taxa_entrega
        total_pago = 0
        
        for i, forma_id in enumerate(formas):
            if forma_id and valores[i]:
                valor = parse_moeda(valores[i])
                if valor > 0:
                    pagamento = PagamentoDelivery(
                        delivery_id=delivery.id,
                        forma_pagamento_id=int(forma_id),
                        valor=valor,
                        observacao=obs_pagamentos[i] if i < len(obs_pagamentos) else ''
                    )
                    db.session.add(pagamento)
                    total_pago += valor
        
        if abs(total_pago - total_com_taxa) > 0.01:
            db.session.rollback()
            flash('O total dos pagamentos não corresponde ao valor total (pedido + taxa)!', 'danger')
            return redirect(url_for('delivery'))
        
        db.session.commit()
        flash('Delivery registrado com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar delivery: {str(e)}', 'danger')
    
    return redirect(url_for('delivery'))

# ==================== ROUTES - DESPESAS ====================

@app.route('/despesas')
@login_required
def despesas():
    caixa = db.session.get(Caixa, session['caixa_id'])
    if not caixa:
        session.clear()
        flash('Caixa não encontrado. Por favor, faça login novamente.', 'warning')
        return redirect(url_for('login'))

    despesas_list = Despesa.query.filter_by(caixa_id=session['caixa_id']).order_by(Despesa.data_hora.desc()).all()
    categorias = CategoriaDespesa.query.filter_by(ativo=True).all()
    formas_pagamento = FormaPagamento.query.filter_by(ativo=True).all()
    
    totais = calcular_totais_caixa(caixa)
    
    return render_template('despesas.html',
                         despesas=despesas_list,
                         categorias=categorias,
                         formas_pagamento=formas_pagamento,
                         totais=totais,
                         caixa=caixa)

@app.route('/despesas/nova', methods=['POST'])
@login_required
def nova_despesa():
    try:
        tipo = request.form.get('tipo')
        categoria_id = request.form.get('categoria_id')
        descricao = request.form.get('descricao')
        valor = parse_moeda(request.form.get('valor'))
        forma_pagamento_id = request.form.get('forma_pagamento_id')
        data_vencimento = request.form.get('data_vencimento')
        observacao = request.form.get('observacao', '')
        
        despesa = Despesa(
            caixa_id=session['caixa_id'],
            tipo=tipo,
            categoria_id=int(categoria_id) if categoria_id else None,
            descricao=descricao,
            valor=valor,
            forma_pagamento_id=int(forma_pagamento_id) if forma_pagamento_id else None,
            data_vencimento=datetime.strptime(data_vencimento, '%Y-%m-%d').date() if data_vencimento else None,
            observacao=observacao
        )
        db.session.add(despesa)
        db.session.commit()
        
        flash('Despesa registrada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar despesa: {str(e)}', 'danger')
    
    return redirect(url_for('despesas'))

# ==================== ROUTES - SANGRIA ====================

@app.route('/sangria')
@login_required
def sangria():
    caixa = db.session.get(Caixa, session['caixa_id'])
    if not caixa:
        session.clear()
        flash('Caixa não encontrado. Por favor, faça login novamente.', 'warning')
        return redirect(url_for('login'))

    sangrias_list = Sangria.query.filter_by(caixa_id=session['caixa_id']).order_by(Sangria.data_hora.desc()).all()
    
    totais = calcular_totais_caixa(caixa)
    
    return render_template('sangria.html',
                         sangrias=sangrias_list,
                         totais=totais,
                         caixa=caixa)

@app.route('/sangria/nova', methods=['POST'])
@login_required
def nova_sangria():
    try:
        valor = parse_moeda(request.form.get('valor'))
        motivo = request.form.get('motivo')
        observacao = request.form.get('observacao', '')
        
        sangria_obj = Sangria(
            caixa_id=session['caixa_id'],
            valor=valor,
            motivo=motivo,
            observacao=observacao
        )
        db.session.add(sangria_obj)
        db.session.commit()
        
        flash('Sangria registrada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar sangria: {str(e)}', 'danger')
    
    return redirect(url_for('sangria'))

# ==================== ROUTES - ESTOQUE ====================

@app.route('/estoque')
@login_required
def estoque():
    produtos = Produto.query.filter_by(ativo=True).all()
    
    # Calcular totais
    total_produtos = len(produtos)
    total_valor = sum(p.quantidade * p.custo for p in produtos)
    criticos = len([p for p in produtos if p.quantidade <= p.estoque_minimo * 0.3])
    baixos = len([p for p in produtos if p.quantidade > p.estoque_minimo * 0.3 and p.quantidade <= p.estoque_minimo])
    
    return render_template('estoque.html',
                         produtos=produtos,
                         total_produtos=total_produtos,
                         total_valor=total_valor,
                         criticos=criticos,
                         baixos=baixos)

@app.route('/estoque/produto/novo', methods=['POST'])
@login_required
def novo_produto():
    try:
        codigo = request.form.get('codigo')
        nome = request.form.get('nome')
        categoria = request.form.get('categoria')
        custo = parse_moeda(request.form.get('custo', 0))
        preco_venda = parse_moeda(request.form.get('preco_venda', 0))
        quantidade = int(request.form.get('quantidade', 0))
        estoque_minimo = int(request.form.get('estoque_minimo', 0))
        estoque_maximo = int(request.form.get('estoque_maximo', 0))
        
        produto = Produto(
            codigo=codigo or f'PROD{Produto.query.count() + 1:03d}',
            nome=nome,
            categoria=categoria,
            custo=custo,
            preco_venda=preco_venda,
            quantidade=quantidade,
            estoque_minimo=estoque_minimo,
            estoque_maximo=estoque_maximo
        )
        db.session.add(produto)
        db.session.flush()
        
        # Criar movimentação inicial se houver quantidade
        if quantidade > 0:
            movimentacao = MovimentacaoEstoque(
                produto_id=produto.id,
                tipo='ENTRADA',
                quantidade=quantidade,
                valor_unitario=custo,
                valor_total=custo * quantidade,
                motivo='CADASTRO',
                observacao='Cadastro inicial',
                usuario_id=session['user_id']
            )
            db.session.add(movimentacao)
        
        db.session.commit()
        flash('Produto cadastrado com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao cadastrar produto: {str(e)}', 'danger')
    
    return redirect(url_for('estoque'))

@app.route('/estoque/movimentacao/nova', methods=['POST'])
@login_required
def nova_movimentacao():
    try:
        produto_id = int(request.form.get('produto_id'))
        tipo = request.form.get('tipo')
        quantidade = int(request.form.get('quantidade'))
        valor_unitario = parse_moeda(request.form.get('valor_unitario', 0))
        motivo = request.form.get('motivo')
        observacao = request.form.get('observacao', '')
        
        produto = db.session.get(Produto, produto_id)
        if not produto:
            flash('Produto não encontrado!', 'danger')
            return redirect(url_for('estoque'))
        
        # Verificar estoque para saídas
        if tipo == 'SAIDA' and quantidade > produto.quantidade:
            flash(f'Quantidade insuficiente! Disponível: {produto.quantidade}', 'danger')
            return redirect(url_for('estoque'))
        
        # Atualizar quantidade
        if tipo == 'ENTRADA':
            produto.quantidade += quantidade
        elif tipo == 'SAIDA':
            produto.quantidade -= quantidade
        elif tipo == 'AJUSTE':
            produto.quantidade = quantidade
        
        # Criar movimentação
        movimentacao = MovimentacaoEstoque(
            produto_id=produto_id,
            tipo=tipo,
            quantidade=quantidade,
            valor_unitario=valor_unitario,
            valor_total=valor_unitario * quantidade,
            motivo=motivo,
            observacao=observacao,
            usuario_id=session['user_id']
        )
        db.session.add(movimentacao)
        db.session.commit()
        
        flash('Movimentação registrada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar movimentação: {str(e)}', 'danger')
    
    return redirect(url_for('estoque'))

# ==================== ROUTES - DASHBOARD ====================

@app.route('/dashboard')
@login_required
@dashboard_required 
def dashboard():
    # Pegar filtros
    periodo = request.args.get('periodo', 'month')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    turno = request.args.get('turno', 'all')
    
    # Definir datas
    hoje = datetime.now().date()
    if periodo == 'today':
        data_inicio = hoje
        data_fim = hoje
    elif periodo == 'week':
        data_inicio = hoje - timedelta(days=7)
        data_fim = hoje
    elif periodo == 'month':
        data_inicio = hoje.replace(day=1)
        data_fim = hoje
    else:
        if data_inicio:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        if data_fim:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    # Buscar caixas no período
    caixas_query = Caixa.query.filter(
        Caixa.data >= data_inicio,
        Caixa.data <= data_fim
    )

    if turno and turno != 'all':
        caixas_query = caixas_query.filter(Caixa.turno == turno)

    caixas = caixas_query.all()
    
    # Calcular métricas
    metricas = calcular_metricas_dashboard(caixas)
    
    # Calcular métricas avançadas
    metricas_avancadas = calcular_metricas_avancadas(caixas)
    
    return render_template('dashboard.html',
                         metricas=metricas,
                         metricas_avancadas=metricas_avancadas,
                         caixas=caixas,
                         periodo=periodo,
                         data_inicio=data_inicio,
                         data_fim=data_fim,
                         turno=turno)

# ==================== ROUTES - CONFIGURAÇÕES ====================

@app.route('/configuracoes')
@admin_required
def configuracoes():
    usuarios = Usuario.query.all()
    formas_pagamento = FormaPagamento.query.all()
    bandeiras = BandeiraCartao.query.all()
    categorias = CategoriaDespesa.query.all()
    motoboys = Motoboy.query.all()
    
    return render_template('configuracoes.html',
                         usuarios=usuarios,
                         formas_pagamento=formas_pagamento,
                         bandeiras=bandeiras,
                         categorias=categorias,
                         motoboys=motoboys)

@app.route('/configuracoes/usuario/novo', methods=['POST'])
@admin_required
def novo_usuario():
    try:
        nome = request.form.get('nome')
        senha = request.form.get('senha')
        perfil = request.form.get('perfil', 'OPERADOR')
        acesso_dashboard = 'acesso_dashboard' in request.form
        acesso_configuracoes = 'acesso_configuracoes' in request.form
        acesso_relatorios = 'acesso_relatorios' in request.form
        
        # Verificar se usuário já existe
        usuario_existe = Usuario.query.filter_by(nome=nome).first()
        if usuario_existe:
            flash('Já existe um usuário com este nome!', 'warning')
            return redirect(url_for('configuracoes'))
        
        usuario = Usuario(
            nome=nome,
            senha=generate_password_hash(senha),
            perfil=perfil,
            acesso_dashboard=acesso_dashboard,
            acesso_configuracoes=acesso_configuracoes,
            acesso_relatorios=acesso_relatorios
        )
        db.session.add(usuario)
        db.session.commit()
        
        flash(f'Usuário {nome} criado com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar usuário: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

@app.route('/configuracoes/forma-pagamento/nova', methods=['POST'])
@admin_required
def nova_forma_pagamento():
    try:
        nome = request.form.get('nome')
        
        if not nome or nome.strip() == '':
            flash('Informe o nome da forma de pagamento!', 'warning')
            return redirect(url_for('configuracoes'))
        
        # Verificar se já existe
        existe = FormaPagamento.query.filter_by(nome=nome).first()
        if existe:
            flash('Esta forma de pagamento já existe!', 'warning')
            return redirect(url_for('configuracoes'))
        
        forma = FormaPagamento(nome=nome)
        db.session.add(forma)
        db.session.commit()
        
        flash(f'Forma de pagamento "{nome}" criada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar forma de pagamento: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

@app.route('/configuracoes/bandeira/nova', methods=['POST'])
@admin_required
def nova_bandeira():
    try:
        nome = request.form.get('nome')
        
        if not nome or nome.strip() == '':
            flash('Informe o nome da bandeira!', 'warning')
            return redirect(url_for('configuracoes'))
        
        # Verificar se já existe
        existe = BandeiraCartao.query.filter_by(nome=nome).first()
        if existe:
            flash('Esta bandeira já existe!', 'warning')
            return redirect(url_for('configuracoes'))
        
        bandeira = BandeiraCartao(nome=nome)
        db.session.add(bandeira)
        db.session.commit()
        
        flash(f'Bandeira "{nome}" criada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar bandeira: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

@app.route('/configuracoes/categoria/nova', methods=['POST'])
@admin_required
def nova_categoria():
    try:
        nome = request.form.get('nome')
        tipo = request.form.get('tipo')
        
        if not nome or nome.strip() == '':
            flash('Informe o nome da categoria!', 'warning')
            return redirect(url_for('configuracoes'))
        
        # Verificar se já existe
        existe = CategoriaDespesa.query.filter_by(nome=nome, tipo=tipo).first()
        if existe:
            flash('Esta categoria já existe!', 'warning')
            return redirect(url_for('configuracoes'))
        
        categoria = CategoriaDespesa(nome=nome, tipo=tipo)
        db.session.add(categoria)
        db.session.commit()
        
        flash(f'Categoria "{nome}" ({tipo}) criada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar categoria: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

@app.route('/configuracoes/motoboy/novo', methods=['POST'])
@admin_required
def novo_motoboy():
    try:
        nome = request.form.get('nome')
        taxa_padrao = parse_moeda(request.form.get('taxa_padrao', 5.00))
        
        if not nome or nome.strip() == '':
            flash('Informe o nome do motoboy!', 'warning')
            return redirect(url_for('configuracoes'))
        
        # Verificar se já existe
        existe = Motoboy.query.filter_by(nome=nome).first()
        if existe:
            flash('Já existe um motoboy com este nome!', 'warning')
            return redirect(url_for('configuracoes'))
        
        motoboy = Motoboy(nome=nome, taxa_padrao=taxa_padrao)
        db.session.add(motoboy)
        db.session.commit()
        
        flash(f'Motoboy "{nome}" cadastrado com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao cadastrar motoboy: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

# ==================== ROUTES - GESTÃO DE CAIXAS (ADMIN) ====================

@app.route('/admin/caixas')
@admin_required
def admin_caixas():
    """Lista todos os caixas para administração"""
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', 'all')
    
    query = Caixa.query.order_by(Caixa.data.desc(), Caixa.hora_abertura.desc())
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter.upper())
    
    caixas = query.paginate(page=page, per_page=20, error_out=False)
    
    return render_template('admin_caixas.html', caixas=caixas, status_filter=status_filter)

@app.route('/admin/caixa/<int:caixa_id>/visualizar')
@admin_required
def admin_visualizar_caixa(caixa_id):
    """Visualizar detalhes de um caixa específico"""
    caixa = db.session.get(Caixa, caixa_id)
    if not caixa:
        flash('Caixa não encontrado!', 'danger')
        return redirect(url_for('admin_caixas'))
    
    totais = calcular_totais_fechamento(caixa)
    
    return render_template('admin_visualizar_caixa.html', caixa=caixa, totais=totais)

@app.route('/admin/caixa/<int:caixa_id>/excluir-completo', methods=['POST'])
@admin_required
def admin_excluir_caixa_completo(caixa_id):
    """Excluir um caixa e TODOS os seus registros (vendas, deliveries, despesas, etc.)"""
    try:
        caixa = db.session.get(Caixa, caixa_id)
        if not caixa:
            flash('Caixa não encontrado!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        # Registrar informações para mensagem
        info_caixa = f"Caixa #{caixa.id} - {caixa.data.strftime('%d/%m/%Y')} - {caixa.turno}"
        
        # Excluir todos os registros relacionados (em ordem para evitar constraints)
        
        # 1. Excluir pagamentos das vendas
        for venda in caixa.vendas:
            PagamentoVenda.query.filter_by(venda_id=venda.id).delete()
        
        # 2. Excluir pagamentos dos deliveries
        for delivery in caixa.deliveries:
            PagamentoDelivery.query.filter_by(delivery_id=delivery.id).delete()
        
        # 3. Excluir vendas
        Venda.query.filter_by(caixa_id=caixa_id).delete()
        
        # 4. Excluir deliveries
        Delivery.query.filter_by(caixa_id=caixa_id).delete()
        
        # 5. Excluir despesas
        Despesa.query.filter_by(caixa_id=caixa_id).delete()
        
        # 6. Excluir sangrias
        Sangria.query.filter_by(caixa_id=caixa_id).delete()
        
        # 7. Excluir suprimentos (se existir)
        if hasattr(caixa, 'suprimentos'):
            Suprimento.query.filter_by(caixa_id=caixa_id).delete()
        
        # 8. Finalmente, excluir o caixa
        db.session.delete(caixa)
        db.session.commit()
        
        flash(f'✅ Caixa {info_caixa} excluído permanentemente com todos os registros!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erro ao excluir caixa: {str(e)}', 'danger')
    
    return redirect(url_for('admin_caixas'))

@app.route('/admin/caixa/<int:caixa_id>/gerar-pdf')
@admin_required
def admin_gerar_pdf_caixa(caixa_id):
    """Gerar PDF do fechamento do caixa"""
    try:
        from flask import make_response
        
        caixa = db.session.get(Caixa, caixa_id)
        if not caixa:
            flash('Caixa não encontrado!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        totais = calcular_totais_fechamento(caixa)
        
        # Renderizar HTML do relatório
        html = render_template('relatorio_imprimivel.html', 
                              caixa=caixa, 
                              totais=totais,
                              now=datetime.now)
        
        # Tentar importar weasyprint para gerar PDF
        try:
            from weasyprint import HTML
            pdf = HTML(string=html).write_pdf()
            
            response = make_response(pdf)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'inline; filename=caixa_{caixa_id}_{caixa.data.strftime("%d%m%Y")}_{caixa.turno}.pdf'
            
            return response
        except Exception as e:
            # Se houver erro (falta de DLLs GTK ou biblioteca não instalada), retorna HTML
            flash(f'Modo de compatibilidade ativado: Imprima usando o navegador (Ctrl+P).', 'warning')
            return html
            
    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'danger')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))

@app.route('/admin/caixa/<int:caixa_id>/editar', methods=['GET', 'POST'])
@admin_required
def admin_editar_caixa(caixa_id):
    """Editar um caixa (apenas admin)"""
    caixa = db.session.get(Caixa, caixa_id)
    if not caixa:
        flash('Caixa não encontrado!', 'danger')
        return redirect(url_for('admin_caixas'))
    
    if request.method == 'POST':
        try:
            caixa.saldo_inicial = parse_moeda(request.form.get('saldo_inicial', caixa.saldo_inicial))
            caixa.saldo_final = parse_moeda(request.form.get('saldo_final', caixa.saldo_final))
            
            db.session.commit()
            flash('Caixa atualizado com sucesso!', 'success')
            return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar caixa: {str(e)}', 'danger')
    
    totais = calcular_totais_fechamento(caixa)
    return render_template('admin_editar_caixa.html', caixa=caixa, totais=totais)

@app.route('/admin/venda/<int:venda_id>/editar', methods=['POST'])
@admin_required
def admin_editar_venda(venda_id):
    """Editar uma venda"""
    try:
        venda = db.session.get(Venda, venda_id)
        if not venda:
            flash('Venda não encontrada!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        venda.total = parse_moeda(request.form.get('total', venda.total))
        venda.observacao = request.form.get('observacao', venda.observacao)
        
        db.session.commit()
        flash('Venda atualizada com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar venda: {str(e)}', 'danger')
    
    return redirect(url_for('admin_visualizar_caixa', caixa_id=venda.caixa_id))

@app.route('/admin/venda/<int:venda_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_venda(venda_id):
    """Deletar uma venda"""
    try:
        venda = db.session.get(Venda, venda_id)
        if not venda:
            flash('Venda não encontrada!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        caixa_id = venda.caixa_id
        
        # Deletar pagamentos relacionados
        PagamentoVenda.query.filter_by(venda_id=venda_id).delete()
        
        # Deletar venda
        db.session.delete(venda)
        db.session.commit()
        
        flash('Venda removida com sucesso!', 'success')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover venda: {str(e)}', 'danger')
        return redirect(url_for('admin_caixas'))

@app.route('/admin/despesa/<int:despesa_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_despesa(despesa_id):
    """Deletar uma despesa"""
    try:
        despesa = db.session.get(Despesa, despesa_id)
        if not despesa:
            flash('Despesa não encontrada!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        caixa_id = despesa.caixa_id
        
        db.session.delete(despesa)
        db.session.commit()
        
        flash('Despesa removida com sucesso!', 'success')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover despesa: {str(e)}', 'danger')
        return redirect(url_for('admin_caixas'))

@app.route('/admin/caixa/<int:caixa_id>/reabrir', methods=['POST'])
@admin_required
def admin_reabrir_caixa(caixa_id):
    """Reabrir um caixa fechado (apenas admin)"""
    try:
        caixa = db.session.get(Caixa, caixa_id)
        if not caixa:
            flash('Caixa não encontrado!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        if caixa.status == 'ABERTO':
            flash('Este caixa já está aberto!', 'warning')
            return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
        
        # Reabrir caixa
        caixa.status = 'ABERTO'
        caixa.hora_fechamento = None
        
        db.session.commit()
        
        flash(f'Caixa #{caixa_id} reaberto com sucesso! Agora pode ser editado.', 'success')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao reabrir caixa: {str(e)}', 'danger')
        return redirect(url_for('admin_caixas'))

@app.route('/admin/caixa/<int:caixa_id>/fechar-forcado', methods=['POST'])
@admin_required
def admin_fechar_caixa_forcado(caixa_id):
    """Fechar um caixa aberto forçadamente (admin)"""
    try:
        caixa = db.session.get(Caixa, caixa_id)
        if not caixa:
            flash('Caixa não encontrado!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        if caixa.status == 'FECHADO':
            flash('Este caixa já está fechado!', 'warning')
            return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
        
        # Calcular saldo final
        totais = calcular_totais_fechamento(caixa)
        caixa.saldo_final = totais['saldo_final']
        caixa.status = 'FECHADO'
        caixa.hora_fechamento = datetime.utcnow()
        
        db.session.commit()
        
        flash(f'⚠️ Caixa #{caixa_id} fechado pelo administrador! Operador: {caixa.operador.nome}', 'warning')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao fechar caixa: {str(e)}', 'danger')
        return redirect(url_for('admin_caixas'))

@app.route('/admin/caixa/<int:caixa_id>/gerar-relatorio')
@admin_required
def admin_gerar_relatorio(caixa_id):
    """Gerar relatório do caixa em formato imprimível"""
    try:
        caixa = db.session.get(Caixa, caixa_id)
        if not caixa:
            flash('Caixa não encontrado!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        totais = calcular_totais_fechamento(caixa)
        
        # Renderizar HTML otimizado para impressão
        return render_template('relatorio_imprimivel.html', 
                             caixa=caixa, 
                             totais=totais,
                             now=datetime.now)
        
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))

# ==================== ROTA EXPORTAR EXCEL REAL (.xlsx) ====================

@app.route('/exportar/excel-real/<int:caixa_id>')
@login_required
def exportar_excel_real(caixa_id):
    """Exportar todos os movimentos do caixa para Excel REAL (.xlsx)"""
    try:
        from io import BytesIO
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        caixa = db.session.get(Caixa, caixa_id)
        if not caixa:
            flash('Caixa não encontrado!', 'danger')
            return redirect(url_for('dashboard'))
        
        # Verificar permissão
        usuario = db.session.get(Usuario, session['user_id'])
        if not usuario.acesso_configuracoes and caixa.operador_id != usuario.id:
            flash('Você não tem permissão para exportar este caixa!', 'danger')
            return redirect(url_for('dashboard'))
        
        # Criar um workbook Excel
        wb = Workbook()
        
        # ========== ESTILOS ==========
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        number_alignment = Alignment(horizontal='right', vertical='center')
        currency_format = '"R$"#,##0.00'
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ========== ABA 1: INFORMAÇÕES GERAIS ==========
        ws_info = wb.active
        ws_info.title = "INFORMAÇÕES"
        
        # Título
        ws_info.merge_cells('A1:D1')
        ws_info['A1'] = f"RELATÓRIO DO CAIXA #{caixa.id}"
        ws_info['A1'].font = Font(name='Arial', size=14, bold=True)
        ws_info['A1'].alignment = Alignment(horizontal='center')
        
        # Informações básicas
        info_data = [
            ['DATA DO CAIXA:', caixa.data.strftime('%d/%m/%Y')],
            ['TURNO:', caixa.turno],
            ['OPERADOR:', caixa.operador.nome],
            ['STATUS:', caixa.status],
            ['HORA ABERTURA:', caixa.hora_abertura.strftime('%H:%M:%S')],
            ['HORA FECHAMENTO:', caixa.hora_fechamento.strftime('%H:%M:%S') if caixa.hora_fechamento else '-'],
            ['SALDO INICIAL:', caixa.saldo_inicial],
            ['SALDO FINAL:', caixa.saldo_final if caixa.saldo_final else '-']
        ]
        
        for i, (label, value) in enumerate(info_data, start=3):
            ws_info[f'A{i}'] = label
            ws_info[f'B{i}'] = value
            ws_info[f'A{i}'].font = Font(name='Arial', size=11, bold=True)
            if isinstance(value, (int, float)):
                ws_info[f'B{i}'].number_format = currency_format
        
        # Ajustar largura das colunas
        ws_info.column_dimensions['A'].width = 25
        ws_info.column_dimensions['B'].width = 30
        
        # ========== ABA 2: VENDAS ==========
        ws_vendas = wb.create_sheet("VENDAS")
        
        # Cabeçalho
        headers_vendas = [
            'ID', 'DATA', 'HORA', 'TIPO', 'NÚMERO', 'VALOR TOTAL',
            'FORMA PAGAMENTO', 'BANDEIRA', 'VALOR PAGO', 'NOTA FISCAL', 'OBSERVAÇÃO'
        ]
        
        for col, header in enumerate(headers_vendas, start=1):
            cell = ws_vendas.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Dados
        row = 2
        for venda in caixa.vendas:
            for pag in venda.pagamentos:
                ws_vendas.cell(row=row, column=1, value=venda.id).alignment = cell_alignment
                ws_vendas.cell(row=row, column=2, value=venda.data_hora.strftime('%d/%m/%Y')).alignment = cell_alignment
                ws_vendas.cell(row=row, column=3, value=venda.data_hora.strftime('%H:%M:%S')).alignment = cell_alignment
                ws_vendas.cell(row=row, column=4, value=venda.tipo).alignment = cell_alignment
                ws_vendas.cell(row=row, column=5, value=venda.numero if venda.numero else '-').alignment = cell_alignment
                ws_vendas.cell(row=row, column=6, value=venda.total).number_format = currency_format
                ws_vendas.cell(row=row, column=6).alignment = number_alignment
                ws_vendas.cell(row=row, column=7, value=pag.forma_pagamento.nome if pag.forma_pagamento else '-').alignment = cell_alignment
                ws_vendas.cell(row=row, column=8, value=pag.bandeira.nome if pag.bandeira else '-').alignment = cell_alignment
                ws_vendas.cell(row=row, column=9, value=pag.valor).number_format = currency_format
                ws_vendas.cell(row=row, column=9).alignment = number_alignment
                ws_vendas.cell(row=row, column=10, value='SIM' if venda.emitiu_nota else 'NÃO').alignment = Alignment(horizontal='center')
                ws_vendas.cell(row=row, column=11, value=venda.observacao if venda.observacao else '-').alignment = cell_alignment
                row += 1
        
        # Ajustar largura das colunas
        column_widths = [8, 12, 10, 10, 8, 12, 15, 12, 12, 12, 30]
        for i, width in enumerate(column_widths, start=1):
            ws_vendas.column_dimensions[get_column_letter(i)].width = width
        
        # Congelar cabeçalho
        ws_vendas.freeze_panes = 'A2'
        
        # Ajustar altura das linhas para melhor visualização
        for row in range(1, ws_vendas.max_row + 1):
            ws_vendas.row_dimensions[row].height = 20
            
        # ========== ABA 3: DELIVERIES ==========
        ws_delivery = wb.create_sheet("DELIVERIES")
        
        headers_delivery = [
            'ID', 'DATA', 'HORA', 'CLIENTE', 'VALOR PEDIDO', 'TAXA ENTREGA',
            'TOTAL', 'MOTOBOY', 'FORMA PAGAMENTO', 'VALOR PAGO', 'NOTA FISCAL', 'OBSERVAÇÃO'
        ]
        
        for col, header in enumerate(headers_delivery, start=1):
            cell = ws_delivery.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        for delivery in caixa.deliveries:
            for pag in delivery.pagamentos:
                ws_delivery.cell(row=row, column=1, value=delivery.id).alignment = cell_alignment
                ws_delivery.cell(row=row, column=2, value=delivery.data_hora.strftime('%d/%m/%Y')).alignment = cell_alignment
                ws_delivery.cell(row=row, column=3, value=delivery.data_hora.strftime('%H:%M:%S')).alignment = cell_alignment
                ws_delivery.cell(row=row, column=4, value=delivery.cliente).alignment = cell_alignment
                ws_delivery.cell(row=row, column=5, value=delivery.total).number_format = currency_format
                ws_delivery.cell(row=row, column=5).alignment = number_alignment
                ws_delivery.cell(row=row, column=6, value=delivery.taxa_entrega).number_format = currency_format
                ws_delivery.cell(row=row, column=6).alignment = number_alignment
                total_com_taxa = delivery.total + delivery.taxa_entrega
                ws_delivery.cell(row=row, column=7, value=total_com_taxa).number_format = currency_format
                ws_delivery.cell(row=row, column=7).alignment = number_alignment
                ws_delivery.cell(row=row, column=8, value=delivery.motoboy.nome if delivery.motoboy else '-').alignment = cell_alignment
                ws_delivery.cell(row=row, column=9, value=pag.forma_pagamento.nome if pag.forma_pagamento else '-').alignment = cell_alignment
                ws_delivery.cell(row=row, column=10, value=pag.valor).number_format = currency_format
                ws_delivery.cell(row=row, column=10).alignment = number_alignment
                ws_delivery.cell(row=row, column=11, value='SIM' if delivery.emitiu_nota else 'NÃO').alignment = Alignment(horizontal='center')
                ws_delivery.cell(row=row, column=12, value=delivery.observacao if delivery.observacao else '-').alignment = cell_alignment
                row += 1
        
        column_widths_delivery = [8, 12, 10, 25, 12, 12, 12, 15, 15, 12, 12, 30]
        for i, width in enumerate(column_widths_delivery, start=1):
            ws_delivery.column_dimensions[get_column_letter(i)].width = width
        
        ws_delivery.freeze_panes = 'A2'
        
        for row in range(1, ws_delivery.max_row + 1):
            ws_delivery.row_dimensions[row].height = 20
        
        # ========== ABA 4: DESPESAS ==========
        ws_despesas = wb.create_sheet("DESPESAS")
        
        headers_despesas = [
            'ID', 'DATA', 'HORA', 'TIPO', 'CATEGORIA', 'DESCRIÇÃO',
            'VALOR', 'FORMA PAGAMENTO', 'OBSERVAÇÃO'
        ]
        
        for col, header in enumerate(headers_despesas, start=1):
            cell = ws_despesas.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        for despesa in caixa.despesas:
            ws_despesas.cell(row=row, column=1, value=despesa.id).alignment = cell_alignment
            ws_despesas.cell(row=row, column=2, value=despesa.data_hora.strftime('%d/%m/%Y')).alignment = cell_alignment
            ws_despesas.cell(row=row, column=3, value=despesa.data_hora.strftime('%H:%M:%S')).alignment = cell_alignment
            ws_despesas.cell(row=row, column=4, value=despesa.tipo).alignment = cell_alignment
            ws_despesas.cell(row=row, column=5, value=despesa.categoria.nome if despesa.categoria else '-').alignment = cell_alignment
            ws_despesas.cell(row=row, column=6, value=despesa.descricao).alignment = cell_alignment
            ws_despesas.cell(row=row, column=7, value=despesa.valor).number_format = currency_format
            ws_despesas.cell(row=row, column=7).alignment = number_alignment
            ws_despesas.cell(row=row, column=8, value=despesa.forma_pagamento.nome if despesa.forma_pagamento else '-').alignment = cell_alignment
            ws_despesas.cell(row=row, column=9, value=despesa.observacao if despesa.observacao else '-').alignment = cell_alignment
            row += 1
        
        column_widths_despesas = [8, 12, 10, 12, 20, 40, 12, 20, 30]
        for i, width in enumerate(column_widths_despesas, start=1):
            ws_despesas.column_dimensions[get_column_letter(i)].width = width
        
        ws_despesas.freeze_panes = 'A2'
        
        for row in range(1, ws_despesas.max_row + 1):
            ws_despesas.row_dimensions[row].height = 20
        
        # ========== ABA 5: SANGRIA ==========
        ws_sangria = wb.create_sheet("SANGRIA")
        
        headers_sangria = [
            'ID', 'DATA', 'HORA', 'MOTIVO', 'VALOR', 'OBSERVAÇÃO'
        ]
        
        for col, header in enumerate(headers_sangria, start=1):
            cell = ws_sangria.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        for sangria in caixa.sangrias:
            ws_sangria.cell(row=row, column=1, value=sangria.id).alignment = cell_alignment
            ws_sangria.cell(row=row, column=2, value=sangria.data_hora.strftime('%d/%m/%Y')).alignment = cell_alignment
            ws_sangria.cell(row=row, column=3, value=sangria.data_hora.strftime('%H:%M:%S')).alignment = cell_alignment
            ws_sangria.cell(row=row, column=4, value=sangria.motivo).alignment = cell_alignment
            ws_sangria.cell(row=row, column=5, value=sangria.valor).number_format = currency_format
            ws_sangria.cell(row=row, column=5).alignment = number_alignment
            ws_sangria.cell(row=row, column=6, value=sangria.observacao if sangria.observacao else '-').alignment = cell_alignment
            row += 1
        
        column_widths_sangria = [8, 12, 10, 30, 12, 40]
        for i, width in enumerate(column_widths_sangria, start=1):
            ws_sangria.column_dimensions[get_column_letter(i)].width = width
        
        ws_sangria.freeze_panes = 'A2'
        
        for row in range(1, ws_sangria.max_row + 1):
            ws_sangria.row_dimensions[row].height = 20
        
        # ========== ABA 6: RESUMO FINANCEIRO ==========
        ws_resumo = wb.create_sheet("RESUMO")
        
        totais = calcular_totais_fechamento(caixa)
        
        resumo_data = [
            ['DESCRIÇÃO', 'VALOR (R$)'],
            ['SALDO INICIAL', caixa.saldo_inicial],
            ['VENDAS MESA/BALCÃO', totais['vendas_loja']],
            ['VENDAS DELIVERY', totais['vendas_delivery']],
            ['TOTAL VENDAS', totais['total_vendas']],
            ['TOTAL DESPESAS', totais['despesas']],
            ['TOTAL SANGRIA', totais['sangrias']],
            ['SALDO FINAL', totais['saldo_final']]
        ]
        
        for r_idx, row_data in enumerate(resumo_data, start=1):
            for c_idx, cell_data in enumerate(row_data, start=1):
                cell = ws_resumo.cell(row=r_idx, column=c_idx, value=cell_data)
                
                if r_idx == 1:  # Cabeçalho
                    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = Font(name='Arial', size=10)
                    if c_idx == 2:  # Coluna de valores
                        cell.number_format = currency_format
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                
                cell.border = thin_border
        
        # Destacar linha do saldo final
        ws_resumo.cell(row=8, column=1).font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        ws_resumo.cell(row=8, column=1).fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        ws_resumo.cell(row=8, column=2).font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        ws_resumo.cell(row=8, column=2).fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        
        ws_resumo.column_dimensions['A'].width = 30
        ws_resumo.column_dimensions['B'].width = 20
        
        # ========== SALVAR EM MEMÓRIA ==========
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Gerar nome do arquivo
        filename = f"caixa_{caixa_id}_{caixa.data.strftime('%d%m%Y')}_{caixa.turno}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Erro ao exportar Excel: {str(e)}', 'danger')
        import traceback
        print(traceback.format_exc())
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))

# ==================== ROUTES - EDIÇÃO DETALHADA (ADMIN) ====================

@app.route('/admin/venda/<int:venda_id>/editar-detalhes', methods=['GET', 'POST'])
@admin_required
def admin_editar_venda_detalhes(venda_id):
    """Editar venda completa com pagamentos"""
    venda = db.session.get(Venda, venda_id)
    if not venda:
        flash('Venda não encontrada!', 'danger')
        return redirect(url_for('admin_caixas'))
    
    if request.method == 'POST':
        try:
            venda.tipo = request.form.get('tipo', venda.tipo)
            venda.numero = request.form.get('numero', venda.numero)
            venda.total = parse_moeda(request.form.get('total', venda.total))
            venda.observacao = request.form.get('observacao', venda.observacao)
            venda.emitiu_nota = 'emitiu_nota' in request.form
            
            # Atualizar pagamentos
            PagamentoVenda.query.filter_by(venda_id=venda_id).delete()
            
            forma_ids = request.form.getlist('forma_pagamento_id[]')
            valores = request.form.getlist('valor[]')
            bandeira_ids = request.form.getlist('bandeira_id[]')
            
            for i in range(len(forma_ids)):
                if forma_ids[i] and valores[i]:
                    pagamento = PagamentoVenda(
                        venda_id=venda_id,
                        forma_pagamento_id=int(forma_ids[i]),
                        valor=parse_moeda(valores[i]),
                        bandeira_id=int(bandeira_ids[i]) if bandeira_ids[i] and bandeira_ids[i] != '' else None
                    )
                    db.session.add(pagamento)
            
            db.session.commit()
            flash('Venda atualizada com sucesso!', 'success')
            return redirect(url_for('admin_visualizar_caixa', caixa_id=venda.caixa_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar venda: {str(e)}', 'danger')
    
    formas_pagamento = FormaPagamento.query.all()
    bandeiras = BandeiraCartao.query.all()
    
    return render_template('admin_editar_venda_detalhes.html', 
                         venda=venda, 
                         formas_pagamento=formas_pagamento,
                         bandeiras=bandeiras)

@app.route('/admin/delivery/<int:delivery_id>/editar-detalhes', methods=['GET', 'POST'])
@admin_required
def admin_editar_delivery_detalhes(delivery_id):
    """Editar delivery completo com pagamentos"""
    delivery = db.session.get(Delivery, delivery_id)
    if not delivery:
        flash('Delivery não encontrado!', 'danger')
        return redirect(url_for('admin_caixas'))
    
    if request.method == 'POST':
        try:
            delivery.cliente = request.form.get('cliente', delivery.cliente)
            if hasattr(delivery, 'endereco'):
                delivery.endereco = request.form.get('endereco', delivery.endereco)
            if hasattr(delivery, 'telefone'):
                delivery.telefone = request.form.get('telefone', delivery.telefone)
            delivery.total = parse_moeda(request.form.get('total', delivery.total))
            delivery.taxa_entrega = parse_moeda(request.form.get('taxa_entrega', delivery.taxa_entrega))
            delivery.observacao = request.form.get('observacao', delivery.observacao)
            delivery.emitiu_nota = 'emitiu_nota' in request.form
            delivery.motoboy_id = int(request.form.get('motoboy_id')) if request.form.get('motoboy_id') else None
            
            # Atualizar pagamentos
            PagamentoDelivery.query.filter_by(delivery_id=delivery_id).delete()
            
            forma_ids = request.form.getlist('forma_pagamento_id[]')
            valores = request.form.getlist('valor[]')
            
            for i in range(len(forma_ids)):
                if forma_ids[i] and valores[i]:
                    pagamento = PagamentoDelivery(
                        delivery_id=delivery_id,
                        forma_pagamento_id=int(forma_ids[i]),
                        valor=parse_moeda(valores[i])
                    )
                    db.session.add(pagamento)
            
            db.session.commit()
            flash('Delivery atualizado com sucesso!', 'success')
            return redirect(url_for('admin_visualizar_caixa', caixa_id=delivery.caixa_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar delivery: {str(e)}', 'danger')
    
    formas_pagamento = FormaPagamento.query.all()
    motoboys = Motoboy.query.all()
    
    return render_template('admin_editar_delivery_detalhes.html', 
                         delivery=delivery,
                         formas_pagamento=formas_pagamento,
                         motoboys=motoboys)

@app.route('/admin/despesa/<int:despesa_id>/editar-detalhes', methods=['GET', 'POST'])
@admin_required
def admin_editar_despesa_detalhes(despesa_id):
    """Editar despesa completa"""
    despesa = db.session.get(Despesa, despesa_id)
    if not despesa:
        flash('Despesa não encontrada!', 'danger')
        return redirect(url_for('admin_caixas'))
    
    if request.method == 'POST':
        try:
            despesa.tipo = request.form.get('tipo', despesa.tipo)
            despesa.descricao = request.form.get('descricao', despesa.descricao)
            despesa.valor = parse_moeda(request.form.get('valor', despesa.valor))
            despesa.categoria_id = int(request.form.get('categoria_id')) if request.form.get('categoria_id') else None
            despesa.forma_pagamento_id = int(request.form.get('forma_pagamento_id')) if request.form.get('forma_pagamento_id') else None
            
            db.session.commit()
            flash('Despesa atualizada com sucesso!', 'success')
            return redirect(url_for('admin_visualizar_caixa', caixa_id=despesa.caixa_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar despesa: {str(e)}', 'danger')
    
    categorias = CategoriaDespesa.query.all()
    formas_pagamento = FormaPagamento.query.all()
    
    return render_template('admin_editar_despesa_detalhes.html',
                         despesa=despesa,
                         categorias=categorias,
                         formas_pagamento=formas_pagamento)

@app.route('/admin/delivery/<int:delivery_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_delivery(delivery_id):
    """Deletar um delivery"""
    try:
        delivery = db.session.get(Delivery, delivery_id)
        if not delivery:
            flash('Delivery não encontrado!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        caixa_id = delivery.caixa_id
        
        # Deletar pagamentos relacionados
        PagamentoDelivery.query.filter_by(delivery_id=delivery_id).delete()
        
        # Deletar delivery
        db.session.delete(delivery)
        db.session.commit()
        
        flash('Delivery removido com sucesso!', 'success')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover delivery: {str(e)}', 'danger')
        return redirect(url_for('admin_caixas'))

# ==================== ROUTES - RELATÓRIOS ====================

@app.route('/relatorios')
@login_required
def relatorios():
    usuario = db.session.get(Usuario, session['user_id'])
    if usuario.perfil in ['ADMIN', 'MASTER']:
        caixas = Caixa.query.order_by(Caixa.data.desc()).all()
    else:
        caixas = Caixa.query.filter_by(
            operador_id=usuario.id
        ).order_by(Caixa.data.desc()).all()

    return render_template(
        'relatorios.html',
        caixas=caixas
    )

@app.route('/relatorios/diario', methods=['GET', 'POST'])
@login_required
def relatorio_diario():
    """Relatório consolidado do dia"""
    if request.method == 'POST':
        data = request.form.get('data')
        data_obj = datetime.strptime(data, '%Y-%m-%d').date()
    else:
        data_obj = datetime.now().date()
        data = data_obj.strftime('%Y-%m-%d')
    
    # Buscar todos os caixas do dia
    caixas_dia = Caixa.query.filter_by(data=data_obj).all()
    
    # Consolidar dados
    relatorio = {
        'data': data_obj,
        'total_caixas': len(caixas_dia),
        'caixas_abertos': sum(1 for c in caixas_dia if c.status == 'ABERTO'),
        'caixas_fechados': sum(1 for c in caixas_dia if c.status == 'FECHADO'),
        'total_vendas': 0,
        'total_despesas': 0,
        'total_sangrias': 0,
        'saldo_dia': 0,
        'caixas': [],  # Lista de caixas com seus totais
        'turnos': {}
    }
    
    for caixa in caixas_dia:
        totais = calcular_totais_caixa(caixa)
        relatorio['total_vendas'] += totais['total_vendas']
        relatorio['total_despesas'] += totais['despesas']
        relatorio['total_sangrias'] += sum(s.valor for s in caixa.sangrias)
        
        # Adicionar caixa com totais
        relatorio['caixas'].append({
            'caixa': caixa,
            'totais': totais
        })
        
        relatorio['turnos'][caixa.turno] = {
            'caixa': caixa,
            'totais': totais,
            'status': caixa.status
        }
    
    relatorio['saldo_dia'] = relatorio['total_vendas'] - relatorio['total_despesas'] - relatorio['total_sangrias']
    
    return render_template('relatorio_diario.html', relatorio=relatorio, data=data)

@app.route('/relatorios/turno/<int:caixa_id>')
@login_required
def relatorio_turno(caixa_id):
    """Relatório detalhado de um turno específico"""
    caixa = db.session.get(Caixa, caixa_id)
    if not caixa:
        flash('Caixa não encontrado!', 'danger')
        return redirect(url_for('relatorios'))
    
    totais = calcular_totais_fechamento(caixa)
    
    return render_template('relatorio_turno.html', caixa=caixa, totais=totais)

# ==================== ROUTES - FECHAR CAIXA ====================

@app.route('/fechar-caixa')
@login_required
def fechar_caixa():
    caixa = db.session.get(Caixa, session['caixa_id'])
    if not caixa:
        session.clear()
        flash('Caixa não encontrado. Por favor, faça login novamente.', 'warning')
        return redirect(url_for('login'))
    
    # Calcular todos os totais
    totais = calcular_totais_fechamento(caixa)
    
    return render_template('fechar_caixa.html', caixa=caixa, totais=totais)

@app.route('/fechar-caixa/confirmar', methods=['POST'])
@login_required
def confirmar_fechamento():
    try:
        caixa = db.session.get(Caixa, session['caixa_id'])
        
        # Calcular saldo final
        totais = calcular_totais_fechamento(caixa)
        caixa.saldo_final = totais['saldo_final']
        caixa.status = 'FECHADO'
        caixa.hora_fechamento = datetime.utcnow()
        
        db.session.commit()
        
        # Limpar sessão
        session.clear()
        
        flash('Caixa fechado com sucesso!', 'success')
        return redirect(url_for('login'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao fechar caixa: {str(e)}', 'danger')
        return redirect(url_for('fechar_caixa'))

# ==================== ROTAS SUPRIMENTOS (v3.0) ====================

@app.route('/suprimentos')
@login_required
def suprimentos():
    """Página de gestão de suprimentos"""
    caixa = db.session.get(Caixa, session.get('caixa_id'))
    if not caixa:
        flash('Caixa não encontrado!', 'danger')
        return redirect(url_for('login'))
    
    suprimentos_list = Suprimento.query.filter_by(caixa_id=session['caixa_id']).order_by(Suprimento.data_hora.desc()).all()
    totais = calcular_totais_caixa(caixa)
    
    return render_template('suprimentos.html', suprimentos=suprimentos_list, totais=totais, caixa=caixa)

@app.route('/suprimento/novo', methods=['POST'])
@login_required
def novo_suprimento():
    """Adicionar novo suprimento"""
    try:
        valor = parse_moeda(request.form.get('valor'))
        motivo = request.form.get('motivo')
        observacao = request.form.get('observacao', '')
        
        suprimento = Suprimento(
            caixa_id=session['caixa_id'],
            valor=valor,
            motivo=motivo,
            observacao=observacao
        )
        db.session.add(suprimento)
        db.session.commit()
        
        flash(f'✅ Suprimento de R$ {valor:.2f} registrado com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erro ao registrar suprimento: {str(e)}', 'danger')
    
    return redirect(url_for('suprimentos'))

@app.route('/admin/suprimento/<int:suprimento_id>/editar', methods=['GET', 'POST'])
@admin_required
def admin_editar_suprimento(suprimento_id):
    """Editar suprimento (somente admin)"""
    suprimento = db.session.get(Suprimento, suprimento_id)
    if not suprimento:
        flash('Suprimento não encontrado!', 'danger')
        return redirect(url_for('suprimentos'))
    
    if request.method == 'POST':
        try:
            suprimento.valor = parse_moeda(request.form.get('valor'))
            suprimento.motivo = request.form.get('motivo')
            suprimento.observacao = request.form.get('observacao', '')
            db.session.commit()
            flash('✅ Suprimento atualizado com sucesso!', 'success')
            return redirect(url_for('suprimentos'))
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Erro ao atualizar: {str(e)}', 'danger')
    
    return render_template('admin_editar_suprimento.html', suprimento=suprimento)

@app.route('/admin/suprimento/<int:suprimento_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_suprimento(suprimento_id):
    """Deletar suprimento (somente admin)"""
    try:
        suprimento = db.session.get(Suprimento, suprimento_id)
        if suprimento:
            db.session.delete(suprimento)
            db.session.commit()
            flash('✅ Suprimento removido com sucesso!', 'success')
        else:
            flash('Suprimento não encontrado!', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erro ao deletar: {str(e)}', 'danger')
    
    return redirect(url_for('suprimentos'))

# ==================== ROTAS EDIÇÃO/EXCLUSÃO ADMIN (v3.0) ====================

@app.route('/admin/sangria/<int:sangria_id>/editar', methods=['GET', 'POST'])
@admin_required
def admin_editar_sangria(sangria_id):
    """Editar sangria (somente admin)"""
    sangria = db.session.get(Sangria, sangria_id)
    if not sangria:
        flash('Sangria não encontrada!', 'danger')
        return redirect(url_for('sangria'))
    
    if request.method == 'POST':
        try:
            sangria.valor = parse_moeda(request.form.get('valor'))
            sangria.motivo = request.form.get('motivo')
            sangria.observacao = request.form.get('observacao', '')
            db.session.commit()
            flash('✅ Sangria atualizada com sucesso!', 'success')
            return redirect(url_for('sangria'))
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Erro ao atualizar: {str(e)}', 'danger')
    
    return render_template('admin_editar_sangria.html', sangria=sangria)

@app.route('/admin/sangria/<int:sangria_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_sangria(sangria_id):
    """Deletar sangria (somente admin)"""
    try:
        sangria = db.session.get(Sangria, sangria_id)
        if sangria:
            db.session.delete(sangria)
            db.session.commit()
            flash('✅ Sangria removida com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erro ao deletar: {str(e)}', 'danger')
    
    return redirect(url_for('sangria'))

@app.route('/admin/venda/<int:venda_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_venda_completa(venda_id):
    """Deletar venda completa (somente admin)"""
    try:
        venda = db.session.get(Venda, venda_id)
        if venda:
            # Deletar pagamentos primeiro
            PagamentoVenda.query.filter_by(venda_id=venda_id).delete()
            db.session.delete(venda)
            db.session.commit()
            flash('✅ Venda removida com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erro ao deletar venda: {str(e)}', 'danger')
    
    return redirect(url_for('vendas'))

@app.route('/admin/delivery/<int:delivery_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_delivery_completo(delivery_id):
    """Deletar delivery completo (somente admin)"""
    try:
        delivery = db.session.get(Delivery, delivery_id)
        if delivery:
            # Deletar pagamentos primeiro
            PagamentoDelivery.query.filter_by(delivery_id=delivery_id).delete()
            db.session.delete(delivery)
            db.session.commit()
            flash('✅ Delivery removido com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erro ao deletar delivery: {str(e)}', 'danger')
    
    return redirect(url_for('delivery'))

@app.route('/admin/despesa/<int:despesa_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_despesa_completa(despesa_id):
    """Deletar despesa completa (somente admin)"""
    try:
        despesa = db.session.get(Despesa, despesa_id)
        if despesa:
            db.session.delete(despesa)
            db.session.commit()
            flash('✅ Despesa removida com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erro ao deletar despesa: {str(e)}', 'danger')
    
    return redirect(url_for('despesas'))

# ==================== ROTAS GESTÃO DE USUÁRIOS (v3.0) ====================

@app.route('/admin/usuarios/editar/<int:usuario_id>', methods=['GET', 'POST'])
@admin_required
def admin_editar_usuario(usuario_id):
    usuario = db.session.get(Usuario, usuario_id)
    if not usuario:
        flash('Usuário não encontrado', 'danger')
        return redirect(url_for('configuracoes'))
    if usuario.perfil == 'MASTER':
        flash('N??o ?? permitido editar o ADMIN MASTER.', 'warning')
        return redirect(url_for('configuracoes'))

    if request.method == 'POST':
        usuario.perfil = request.form.get('perfil')
        
        # Para checkboxes, usamos 'in' para verificar se foram enviados
        usuario.acesso_dashboard = 'acesso_dashboard' in request.form
        usuario.acesso_configuracoes = 'acesso_configuracoes' in request.form
        usuario.acesso_relatorios = 'acesso_relatorios' in request.form
        
        db.session.commit()
        flash('Usuário atualizado com sucesso', 'success')
        return redirect(url_for('configuracoes'))

    return render_template(
        'editar_usuario.html',
        usuario=usuario
    )

@app.route('/admin/usuarios/toggle/<int:usuario_id>')
@admin_required
def admin_toggle_usuario(usuario_id):
    usuario = db.session.get(Usuario, usuario_id)
    if usuario and usuario.perfil == 'MASTER':
        flash('N??o ?? permitido alterar o ADMIN MASTER.', 'warning')
        return redirect(url_for('configuracoes'))
    if usuario and usuario.perfil not in ['ADMIN', 'MASTER']:
        usuario.ativo = not usuario.ativo
        db.session.commit()
        flash('Status do usuário alterado', 'info')
    return redirect(url_for('configuracoes'))

@app.route('/admin/usuario/<int:usuario_id>/editar-senha', methods=['POST'])
@admin_required
def admin_editar_senha_usuario(usuario_id):
    """Alterar senha de usuário (somente admin)"""
    try:
        usuario = db.session.get(Usuario, usuario_id)
        if not usuario:
            flash('Usuário não encontrado!', 'danger')
            return redirect(url_for('configuracoes'))
        if usuario.perfil == 'MASTER':
            flash('N??o ?? permitido alterar a senha do ADMIN MASTER.', 'warning')
            return redirect(url_for('configuracoes'))
        
        nova_senha = request.form.get('nova_senha')
        if nova_senha and len(nova_senha) >= 3:
            usuario.senha = generate_password_hash(nova_senha)
            db.session.commit()
            flash(f'✅ Senha de {usuario.nome} alterada com sucesso!', 'success')
        else:
            flash('❌ Senha deve ter no mínimo 3 caracteres!', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erro ao alterar senha: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

@app.route('/admin/usuario/<int:usuario_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_usuario(usuario_id):
    """Deletar usuário (somente admin)"""
    try:
        usuario = db.session.get(Usuario, usuario_id)
        if not usuario:
            flash('Usuário não encontrado!', 'danger')
            return redirect(url_for('configuracoes'))
        if usuario.perfil == 'MASTER':
            flash('N??o ?? permitido excluir o ADMIN MASTER.', 'warning')
            return redirect(url_for('configuracoes'))
        
        # Verificar se não é o último admin
        if usuario.acesso_configuracoes:
            total_admins = Usuario.query.filter_by(acesso_configuracoes=True, ativo=True).count()
            if total_admins <= 1:
                flash('❌ Não é possível excluir o último administrador do sistema!', 'danger')
                return redirect(url_for('configuracoes'))
        
        # Verificar se tem caixas vinculados
        if len(usuario.caixas) > 0:
            flash('❌ Não é possível excluir usuário com caixas registrados! Primeiro transfira ou remova os caixas.', 'danger')
            return redirect(url_for('configuracoes'))
        
        nome_usuario = usuario.nome
        db.session.delete(usuario)
        db.session.commit()
        flash(f'✅ Usuário {nome_usuario} excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Erro ao deletar usuário: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

# ==================== ROTAS PARA FORMAS DE PAGAMENTO ====================

@app.route('/admin/forma-pagamento/<int:forma_id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_forma_pagamento(forma_id):
    """Ativar/Desativar forma de pagamento"""
    forma = db.session.get(FormaPagamento, forma_id)
    if forma:
        forma.ativo = not forma.ativo
        db.session.commit()
        flash(f'Forma de pagamento "{forma.nome}" atualizada!', 'success')
    return redirect(url_for('configuracoes'))

@app.route('/admin/forma-pagamento/<int:forma_id>/editar', methods=['POST'])
@admin_required
def admin_editar_forma_pagamento(forma_id):
    """Editar forma de pagamento"""
    try:
        forma = db.session.get(FormaPagamento, forma_id)
        if forma:
            forma.nome = request.form.get('nome')
            forma.ativo = 'ativo' in request.form
            db.session.commit()
            flash('Forma de pagamento atualizada!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

@app.route('/admin/forma-pagamento/<int:forma_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_forma_pagamento(forma_id):
    """Deletar forma de pagamento"""
    try:
        forma = db.session.get(FormaPagamento, forma_id)
        if forma:
            db.session.delete(forma)
            db.session.commit()
            flash('Forma de pagamento excluída!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

# ==================== ROTAS PARA BANDEIRAS ====================

@app.route('/admin/bandeira/<int:bandeira_id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_bandeira(bandeira_id):
    """Ativar/Desativar bandeira"""
    bandeira = db.session.get(BandeiraCartao, bandeira_id)
    if bandeira:
        bandeira.ativo = not bandeira.ativo
        db.session.commit()
        flash(f'Bandeira "{bandeira.nome}" atualizada!', 'success')
    return redirect(url_for('configuracoes'))

@app.route('/admin/bandeira/<int:bandeira_id>/editar', methods=['POST'])
@admin_required
def admin_editar_bandeira(bandeira_id):
    """Editar bandeira"""
    try:
        bandeira = db.session.get(BandeiraCartao, bandeira_id)
        if bandeira:
            bandeira.nome = request.form.get('nome')
            bandeira.ativo = 'ativo' in request.form
            db.session.commit()
            flash('Bandeira atualizada!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

@app.route('/admin/bandeira/<int:bandeira_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_bandeira(bandeira_id):
    """Deletar bandeira"""
    try:
        bandeira = db.session.get(BandeiraCartao, bandeira_id)
        if bandeira:
            db.session.delete(bandeira)
            db.session.commit()
            flash('Bandeira excluída!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

# ==================== ROTAS PARA CATEGORIAS ====================

@app.route('/admin/categoria/<int:categoria_id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_categoria(categoria_id):
    """Ativar/Desativar categoria"""
    categoria = db.session.get(CategoriaDespesa, categoria_id)
    if categoria:
        categoria.ativo = not categoria.ativo
        db.session.commit()
        flash(f'Categoria "{categoria.nome}" atualizada!', 'success')
    return redirect(url_for('configuracoes'))

@app.route('/admin/categoria/<int:categoria_id>/editar', methods=['POST'])
@admin_required
def admin_editar_categoria(categoria_id):
    """Editar categoria"""
    try:
        categoria = db.session.get(CategoriaDespesa, categoria_id)
        if categoria:
            categoria.nome = request.form.get('nome')
            categoria.tipo = request.form.get('tipo')
            categoria.ativo = 'ativo' in request.form
            db.session.commit()
            flash('Categoria atualizada!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

@app.route('/admin/categoria/<int:categoria_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_categoria(categoria_id):
    """Deletar categoria"""
    try:
        categoria = db.session.get(CategoriaDespesa, categoria_id)
        if categoria:
            db.session.delete(categoria)
            db.session.commit()
            flash('Categoria excluída!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

# ==================== ROTAS PARA MOTOBOYS ====================

@app.route('/admin/motoboy/<int:motoboy_id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_motoboy(motoboy_id):
    """Ativar/Desativar motoboy"""
    motoboy = db.session.get(Motoboy, motoboy_id)
    if motoboy:
        motoboy.ativo = not motoboy.ativo
        db.session.commit()
        flash(f'Motoboy "{motoboy.nome}" atualizado!', 'success')
    return redirect(url_for('configuracoes'))

@app.route('/admin/motoboy/<int:motoboy_id>/editar', methods=['POST'])
@admin_required
def admin_editar_motoboy(motoboy_id):
    """Editar motoboy"""
    try:
        motoboy = db.session.get(Motoboy, motoboy_id)
        if motoboy:
            motoboy.nome = request.form.get('nome')
            motoboy.taxa_padrao = parse_moeda(request.form.get('taxa_padrao', 5.00))
            motoboy.ativo = 'ativo' in request.form
            db.session.commit()
            flash('Motoboy atualizado!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

@app.route('/admin/motoboy/<int:motoboy_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_motoboy(motoboy_id):
    """Deletar motoboy"""
    try:
        motoboy = db.session.get(Motoboy, motoboy_id)
        if motoboy:
            db.session.delete(motoboy)
            db.session.commit()
            flash('Motoboy excluído!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

@app.route('/admin/licenca/<int:licenca_id>/toggle', methods=['POST'])
def toggle_licenca(licenca_id):
    """Ativar/Desativar licença"""
    licenca = db.session.get(Licenca, licenca_id)
    if licenca:
        licenca.ativo = not licenca.ativo
        db.session.commit()
        flash(f'Licença {licenca.email} {"ativada" if licenca.ativo else "desativada"}!', 'success')
    return redirect(url_for('admin_licencas'))

@app.route('/admin/licenca/<int:licenca_id>/deletar', methods=['POST'])
def deletar_licenca(licenca_id):
    """Deletar licença"""
    try:
        licenca = db.session.get(Licenca, licenca_id)
        if licenca:
            # Deletar dispositivos associados primeiro
            Dispositivo.query.filter_by(licenca_id=licenca_id).delete()
            # Deletar licença
            db.session.delete(licenca)
            db.session.commit()
            flash(f'Licença {licenca.email} excluída com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir licença: {str(e)}', 'danger')
    return redirect(url_for('admin_licencas'))



#+++++++++++++++++++++teste
@app.route('/debug/licencas')
def debug_licencas():
    """Debug: Mostrar licenças no terminal"""
    licencas = Licenca.query.all()
    print("\n" + "="*50)
    print("LICENÇAS CADASTRADAS NO SISTEMA:")
    print("="*50)
    
    for licenca in licencas:
        print(f"\nID: {licenca.id}")
        print(f"E-mail: {licenca.email}")
        print(f"Chave: {licenca.chave_ativacao}")
        print(f"Status: {licenca.status}")
        print(f"Ativação: {licenca.data_ativacao}")
        print(f"Expiração: {licenca.data_expiracao}")
        print(f"Ativa: {'SIM' if licenca.ativo else 'NÃO'}")
        print(f"Dispositivos: {len(licenca.dispositivos)}/{licenca.max_dispositivos}")
    
    print("\n" + "="*50)
    return "Verifique o terminal do Python para ver as licenças cadastradas."



# ==================== ROTA EXPORTAR TODOS CAIXAS ====================

@app.route('/exportar/todos-caixas')
@admin_required
def exportar_todos_caixas():
    """Exportar todos os caixas para CSV"""
    import csv
    from io import StringIO
    from flask import Response
    
    output = StringIO()
    output.write('\ufeff')  # BOM UTF-8 para Excel PT-BR
    writer = csv.writer(output, delimiter=',')

    
    # Cabeçalho
    writer.writerow([
        'ID', 'Data', 'Turno', 'Operador', 'Status', 'Saldo Inicial', 'Saldo Final',
        'Total Vendas', 'Total Despesas', 'Total Sangrias', 'Hora Abertura', 'Hora Fechamento'
    ])
    
    # Buscar todos os caixas
    caixas = Caixa.query.order_by(Caixa.data.desc(), Caixa.turno).all()
    
    for caixa in caixas:
        totais = calcular_totais_caixa(caixa)
        
        writer.writerow([
            caixa.id,
            caixa.data.strftime('%d/%m/%Y'),
            caixa.turno,
            caixa.operador.nome,
            caixa.status,
            f'{caixa.saldo_inicial:.2f}',
            f'{caixa.saldo_final:.2f}' if caixa.saldo_final else '',
            f'{totais["total_vendas"]:.2f}',
            f'{totais["despesas"]:.2f}',
            f'{totais["sangrias"]:.2f}',
            caixa.hora_abertura.strftime('%H:%M:%S'),
            caixa.hora_fechamento.strftime('%H:%M:%S') if caixa.hora_fechamento else ''
        ])
    
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': 'attachment; filename=todos_caixas.csv'
        }
    )

# ==================== ROTA EXPORTAR EXCEL (v3.0) ====================

@app.route('/exportar/excel/<int:caixa_id>')
@login_required
def exportar_excel_caixa(caixa_id):
    """Exportar todos os movimentos do caixa para Excel/CSV"""
    import csv
    from io import StringIO
    from flask import Response
    
    caixa = db.session.get(Caixa, caixa_id)
    if not caixa:
        flash('Caixa não encontrado!', 'danger')
        return redirect(url_for('dashboard'))
    
    # Verificar permissão (operador só pode exportar próprio caixa)
    usuario = db.session.get(Usuario, session['user_id'])
    if not usuario.acesso_configuracoes and caixa.operador_id != usuario.id:
        flash('Você não tem permissão para exportar este caixa!', 'danger')
        return redirect(url_for('dashboard'))
    
    output = StringIO()
    output.write('\ufeff')  # BOM UTF-8 (ESSENCIAL)
    writer = csv.writer(output, delimiter=',')
    
    # Cabeçalho com TODAS as colunas solicitadas
    writer.writerow([
        'ID',
        'Data',
        'Hora',
        'Turno',
        'Operador',
        'Tipo Movimento',
        'Tipo Venda',
        'Número Mesa/Balcão',
        'Cliente',
        'Endereço',
        'Telefone',
        'Valor Bruto',
        'Valor Líquido',
        'Forma Pagamento',
        'Bandeira',
        'Taxa Entrega',
        'Motoboy',
        'Categoria Despesa',
        'Descrição',
        'Observações',
        'Nota Fiscal'
    ])
    
    # VENDAS
    for venda in caixa.vendas:
        for pag in venda.pagamentos:
            writer.writerow([
                venda.id,
                venda.data_hora.strftime('%d/%m/%Y'),
                venda.data_hora.strftime('%H:%M:%S'),
                caixa.turno,
                caixa.operador.nome,
                'VENDA',
                venda.tipo,
                venda.numero if venda.numero else '-',
                '-',
                '-',
                '-',
                f'{pag.valor:.2f}',
                f'{pag.valor:.2f}',
                pag.forma_pagamento.nome if pag.forma_pagamento else '-',
                pag.bandeira.nome if pag.bandeira else '-',
                '-',
                '-',
                '-',
                '-',
                venda.observacao if venda.observacao else '-',
                'Sim' if venda.emitiu_nota else 'Não'
            ])
    
    # DELIVERIES
    for delivery in caixa.deliveries:
        total_bruto = delivery.total + delivery.taxa_entrega
        for pag in delivery.pagamentos:
            bandeira_nome = '-'
            if hasattr(pag, 'bandeira') and pag.bandeira:
                bandeira_nome = pag.bandeira.nome
            
            writer.writerow([
                delivery.id,
                delivery.data_hora.strftime('%d/%m/%Y'),
                delivery.data_hora.strftime('%H:%M:%S'),
                caixa.turno,
                caixa.operador.nome,
                'DELIVERY',
                '-',
                '-',
                delivery.cliente,
                getattr(delivery, 'endereco', '-') or '-',
                getattr(delivery, 'telefone', '-') or '-',
                f'{total_bruto:.2f}',
                f'{pag.valor:.2f}',
                pag.forma_pagamento.nome if pag.forma_pagamento else '-',
                bandeira_nome,
                f'{delivery.taxa_entrega:.2f}',
                delivery.motoboy.nome if delivery.motoboy else '-',
                '-',
                '-',
                delivery.observacao if delivery.observacao else '-',
                'Sim' if delivery.emitiu_nota else 'Não'
            ])
    
    # DESPESAS
    for despesa in caixa.despesas:
        writer.writerow([
            despesa.id,
            despesa.data_hora.strftime('%d/%m/%Y'),
            despesa.data_hora.strftime('%H:%M:%S'),
            caixa.turno,
            caixa.operador.nome,
            'DESPESA',
            despesa.tipo,
            '-',
            '-',
            '-',
            '-',
            f'{despesa.valor:.2f}',
            f'{despesa.valor:.2f}',
            despesa.forma_pagamento.nome if despesa.forma_pagamento else '-',
            '-',
            '-',
            '-',
            despesa.categoria.nome if despesa.categoria else '-',
            despesa.descricao,
            despesa.observacao if despesa.observacao else '-',
            '-'
        ])
    
    # SANGRIAS
    for sangria in caixa.sangrias:
        writer.writerow([
            sangria.id,
            sangria.data_hora.strftime('%d/%m/%Y'),
            sangria.data_hora.strftime('%H:%M:%S'),
            caixa.turno,
            caixa.operador.nome,
            'SANGRIA',
            '-',
            '-',
            '-',
            '-',
            '-',
            f'{sangria.valor:.2f}',
            f'-{sangria.valor:.2f}',
            '-',
            '-',
            '-',
            '-',
            '-',
            sangria.motivo,
            sangria.observacao if sangria.observacao else '-',
            '-'
        ])
    
    # SUPRIMENTOS
    try:
        if hasattr(caixa, 'suprimentos'):
            for suprimento in caixa.suprimentos:
                writer.writerow([
                    suprimento.id,
                    suprimento.data_hora.strftime('%d/%m/%Y'),
                    suprimento.data_hora.strftime('%H:%M:%S'),
                    caixa.turno,
                    caixa.operador.nome,
                    'SUPRIMENTO',
                    '-',
                    '-',
                    '-',
                    '-',
                    '-',
                    f'{suprimento.valor:.2f}',
                    f'{suprimento.valor:.2f}',
                    '-',
                    '-',
                    '-',
                    '-',
                    '-',
                    suprimento.motivo,
                    suprimento.observacao if suprimento.observacao else '-',
                    '-'
                ])
    except:
        pass  # Caso tabela suprimento não exista ainda
    
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': 'attachment; filename=todos_caixas.csv'}
    )

# ==================== HELPER FUNCTIONS ====================

def calcular_totais_caixa(caixa):
    """Calcula todos os totais do caixa"""
    totais = {
        'vendas_loja': 0,
        'vendas_delivery': 0,
        'total_vendas': 0,
        'dinheiro': 0,
        'credito': 0,
        'debito': 0,
        'pix': 0,
        'online': 0,
        'notas_fiscais': 0,
        'despesas': 0,
        'sangrias': 0,
        'saldo_atual': caixa.saldo_inicial
    }
    
    # Vendas
    for venda in caixa.vendas:
        totais['vendas_loja'] += venda.total
        if venda.emitiu_nota:
            totais['notas_fiscais'] += venda.total
        
        for pagamento in venda.pagamentos:
            if not pagamento.forma_pagamento:
                # Forma de pagamento removida do cadastro
                continue
            forma = pagamento.forma_pagamento.nome.upper()
            if 'DINHEIRO' in forma:
                totais['dinheiro'] += pagamento.valor
            elif 'CRÉDITO' in forma or 'CREDITO' in forma:
                totais['credito'] += pagamento.valor
            elif 'DÉBITO' in forma or 'DEBITO' in forma:
                totais['debito'] += pagamento.valor
            elif 'PIX' in forma:
                totais['pix'] += pagamento.valor
            elif 'ONLINE' in forma:
                totais['online'] += pagamento.valor
    
    # Delivery
    for delivery in caixa.deliveries:
        totais['vendas_delivery'] += delivery.total + delivery.taxa_entrega
        if delivery.emitiu_nota:
            totais['notas_fiscais'] += delivery.total + delivery.taxa_entrega
        
        for pagamento in delivery.pagamentos:
            if not pagamento.forma_pagamento:
                # Forma de pagamento removida do cadastro
                continue
            forma = pagamento.forma_pagamento.nome.upper()
            if 'DINHEIRO' in forma:
                totais['dinheiro'] += pagamento.valor
            elif 'CRÉDITO' in forma or 'CREDITO' in forma:
                totais['credito'] += pagamento.valor
            elif 'DÉBITO' in forma or 'DEBITO' in forma:
                totais['debito'] += pagamento.valor
            elif 'PIX' in forma:
                totais['pix'] += pagamento.valor
            elif 'ONLINE' in forma:
                totais['online'] += pagamento.valor
    
    # Despesas
    for despesa in caixa.despesas:
        totais['despesas'] += despesa.valor
    
    # Sangrias
    for sangria in caixa.sangrias:
        totais['sangrias'] += sangria.valor
    
    totais['total_vendas'] = totais['vendas_loja'] + totais['vendas_delivery']
    totais['saldo_atual'] = caixa.saldo_inicial + totais['total_vendas'] + totais.get('suprimentos', 0) - totais['despesas'] - totais['sangrias']
    
    return totais

def calcular_totais_delivery(caixa):
    """Calcula totais específicos do delivery"""
    totais = {
        'total_delivery': 0,
        'total_taxas': 0,
        'quantidade_pedidos': 0,
        'motoboys': {}
    }
    
    for delivery in caixa.deliveries:
        totais['total_delivery'] += delivery.total + delivery.taxa_entrega
        totais['total_taxas'] += delivery.taxa_entrega
        totais['quantidade_pedidos'] += 1
        
        if delivery.motoboy:
            nome = delivery.motoboy.nome
            if nome not in totais['motoboys']:
                totais['motoboys'][nome] = 0
            totais['motoboys'][nome] += delivery.taxa_entrega
    
    return totais

def calcular_totais_fechamento(caixa):
    """Calcula todos os totais para fechamento"""
    totais = calcular_totais_caixa(caixa)
    
    # Adicionar informações extras para fechamento
    totais['despesas_fixas'] = sum(d.valor for d in caixa.despesas if d.tipo == 'FIXA')
    totais['despesas_variaveis'] = sum(d.valor for d in caixa.despesas if d.tipo == 'VARIAVEL')
    totais['despesas_saidas'] = sum(d.valor for d in caixa.despesas if d.tipo == 'SAIDA')
    totais['saldo_final'] = totais['saldo_atual']
    
    return totais

def calcular_metricas_dashboard(caixas):
    """Calcula métricas para o dashboard - CORRIGIDO"""
    metricas = {
        'total_receitas': 0,
        'total_despesas': 0,
        'saldo_liquido': 0,
        'ticket_medio': 0,
        'total_transacoes': 0,
        'formas_pagamento': {},
        'tipos_venda': {'MESA': 0, 'BALCAO': 0, 'DELIVERY': 0},
        'despesas_categoria': {},
        'vendas_count': 0,
        'delivery_count': 0
    }
    
    for caixa in caixas:
        totais = calcular_totais_caixa(caixa)
        metricas['total_receitas'] += totais.get('total_vendas', 0)
        metricas['total_despesas'] += totais.get('despesas', 0)
        
        # Contar transações
        metricas['vendas_count'] += len(caixa.vendas)
        metricas['delivery_count'] += len(caixa.deliveries)
        metricas['total_transacoes'] = metricas['vendas_count'] + metricas['delivery_count']
        
        # Formas de pagamento
        # Vendas
        for venda in caixa.vendas:
            for pagamento in venda.pagamentos:
                if not pagamento.forma_pagamento:
                    continue
                forma = pagamento.forma_pagamento.nome
                metricas['formas_pagamento'][forma] = metricas['formas_pagamento'].get(forma, 0) + pagamento.valor
            
            # Tipos de venda
            metricas['tipos_venda'][venda.tipo] += venda.total
        
        # Deliveries
        for delivery in caixa.deliveries:
            for pagamento in delivery.pagamentos:
                if not pagamento.forma_pagamento:
                    continue
                forma = pagamento.forma_pagamento.nome
                metricas['formas_pagamento'][forma] = metricas['formas_pagamento'].get(forma, 0) + pagamento.valor
            
            metricas['tipos_venda']['DELIVERY'] += delivery.total + delivery.taxa_entrega
        
        # Despesas por categoria
        for despesa in caixa.despesas:
            if despesa.categoria:
                cat = despesa.categoria.nome
                metricas['despesas_categoria'][cat] = metricas['despesas_categoria'].get(cat, 0) + despesa.valor
    
    metricas['saldo_liquido'] = metricas['total_receitas'] - metricas['total_despesas']
    metricas['ticket_medio'] = metricas['total_receitas'] / metricas['total_transacoes'] if metricas['total_transacoes'] > 0 else 0
    try:
        metricas['ticket_medio'] = round(float(metricas['ticket_medio']), 2)
    except Exception:
        metricas['ticket_medio'] = 0
    
    return metricas

def calcular_metricas_avancadas(caixas):
    """Calcula métricas avançadas para dashboard - CORRIGIDO"""
    metricas = {
        'vendas_por_turno': {'MANHÃ': 0, 'TARDE': 0, 'NOITE': 0},
        'transacoes_por_turno': {'MANHÃ': 0, 'TARDE': 0, 'NOITE': 0},
        'motoboys_taxas': {},
        'despesas_por_tipo': {'FIXA': 0, 'VARIAVEL': 0, 'SAIDA': 0},
        'contas_assinadas': 0,
        'total_sangrias': 0,
        'margem_lucro': 0,
        'custo_operacional': 0,
        'lucratividade': 0,
        'vendas_por_dia': {},
        'despesas_por_dia': {},
        'melhor_dia': {'dia': '-', 'valor': 0},
        'pior_dia': {'dia': '-', 'valor': 99999999},
        'total_notas_fiscais': 0,
        'percentual_notas': 0,
        'ticket_medio_mesa': 0,
        'ticket_medio_delivery': 0,
        'total_produtos_vendidos': 0,
        'vendas_mesa_count': 0,
        'vendas_balcao_count': 0,
        'vendas_delivery_count': 0
    }
    
    total_vendas_mesa = 0
    count_vendas_mesa = 0
    total_vendas_balcao = 0
    count_vendas_balcao = 0
    total_vendas_delivery = 0
    count_vendas_delivery = 0

    def normalizar_turno(turno):
        if not turno:
            return 'MANHÃ'
        t = str(turno).upper()
        if 'MANH' in t:
            return 'MANHÃ'
        if 'TARDE' in t:
            return 'TARDE'
        if 'NOITE' in t:
            return 'NOITE'
        return 'MANHÃ'
    
    for caixa in caixas:
        turno = normalizar_turno(caixa.turno)
        dia_str = caixa.data.strftime('%d/%m')
        
        # Inicializar dia se não existe
        if dia_str not in metricas['vendas_por_dia']:
            metricas['vendas_por_dia'][dia_str] = 0
        if dia_str not in metricas['despesas_por_dia']:
            metricas['despesas_por_dia'][dia_str] = 0
        
        # Vendas
        for venda in caixa.vendas:
            metricas['vendas_por_turno'][turno] += venda.total
            metricas['transacoes_por_turno'][turno] += 1
            metricas['vendas_por_dia'][dia_str] += venda.total
            
            if venda.emitiu_nota:
                metricas['total_notas_fiscais'] += venda.total
            
            # Contar por tipo
            if venda.tipo == 'MESA':
                total_vendas_mesa += venda.total
                count_vendas_mesa += 1
                metricas['vendas_mesa_count'] += 1
            elif venda.tipo == 'BALCAO':
                total_vendas_balcao += venda.total
                count_vendas_balcao += 1
                metricas['vendas_balcao_count'] += 1
            
            # Contar contas assinadas
            for pagamento in venda.pagamentos:
                if not pagamento.forma_pagamento:
                    continue
                if 'ASSINADA' in pagamento.forma_pagamento.nome.upper() or 'CONTA' in pagamento.forma_pagamento.nome.upper():
                    metricas['contas_assinadas'] += pagamento.valor
        
        # Delivery
        for delivery in caixa.deliveries:
            total = delivery.total + delivery.taxa_entrega
            metricas['vendas_por_turno'][turno] += total
            metricas['transacoes_por_turno'][turno] += 1
            metricas['vendas_por_dia'][dia_str] += total
            metricas['vendas_delivery_count'] += 1
            
            if delivery.emitiu_nota:
                metricas['total_notas_fiscais'] += total
            
            # Ticket médio delivery
            total_vendas_delivery += total
            count_vendas_delivery += 1
            
            # Taxas por motoboy
            if delivery.motoboy:
                nome = delivery.motoboy.nome
                if nome not in metricas['motoboys_taxas']:
                    metricas['motoboys_taxas'][nome] = {'total': 0, 'quantidade': 0}
                metricas['motoboys_taxas'][nome]['total'] += delivery.taxa_entrega
                metricas['motoboys_taxas'][nome]['quantidade'] += 1
            
            # Contar contas assinadas em delivery
            for pagamento in delivery.pagamentos:
                if not pagamento.forma_pagamento:
                    continue
                if 'ASSINADA' in pagamento.forma_pagamento.nome.upper() or 'CONTA' in pagamento.forma_pagamento.nome.upper():
                    metricas['contas_assinadas'] += pagamento.valor
        
        # Despesas por tipo
        for despesa in caixa.despesas:
            if despesa.tipo in metricas['despesas_por_tipo']:
                metricas['despesas_por_tipo'][despesa.tipo] += despesa.valor
        
        # Sangrias
        for sangria in caixa.sangrias:
            metricas['total_sangrias'] += sangria.valor
    
    # Calcular melhor e pior dia
    if metricas['vendas_por_dia']:
        for dia, valor in metricas['vendas_por_dia'].items():
            if valor > metricas['melhor_dia']['valor']:
                metricas['melhor_dia'] = {'dia': dia, 'valor': valor}
            if valor < metricas['pior_dia']['valor']:
                metricas['pior_dia'] = {'dia': dia, 'valor': valor}
    
    # Calcular tickets médios
    metricas['ticket_medio_mesa'] = total_vendas_mesa / count_vendas_mesa if count_vendas_mesa > 0 else 0
    metricas['ticket_medio_delivery'] = total_vendas_delivery / count_vendas_delivery if count_vendas_delivery > 0 else 0
    
    # Calcular métricas financeiras
    total_receitas = sum(metricas['vendas_por_turno'].values())
    total_despesas = sum(metricas['despesas_por_tipo'].values())
    
    metricas['custo_operacional'] = total_despesas
    metricas['margem_lucro'] = ((total_receitas - total_despesas) / total_receitas * 100) if total_receitas > 0 else 0
    metricas['lucratividade'] = total_receitas - total_despesas
    metricas['percentual_notas'] = (metricas['total_notas_fiscais'] / total_receitas * 100) if total_receitas > 0 else 0
    
    return metricas

# ==================== INITIALIZE DATABASE ====================

def init_db():
    """Inicializa o banco de dados com dados padrão"""
    with app.app_context():
        db.create_all()
        
        # Verificar se ja existem dados
        if Usuario.query.first():
            master = Usuario.query.filter_by(nome='ADMIN MASTER').first()
            if master:
                if master.perfil != 'MASTER':
                    master.perfil = 'MASTER'
                master.acesso_dashboard = True
                master.acesso_configuracoes = True
                master.acesso_relatorios = True
                master.ativo = True
                db.session.commit()
                return
            master = Usuario.query.filter_by(perfil='MASTER').first()
            if not master:
                master = Usuario(
                    nome='ADMIN MASTER',
                    senha=generate_password_hash('SUPORTE26@'),
                    perfil='MASTER',
                    acesso_dashboard=True,
                    acesso_configuracoes=True,
                    acesso_relatorios=True
                )
                db.session.add(master)
                db.session.commit()
            return
        
        # Criar usuario admin master
        master = Usuario(
            nome='ADMIN MASTER',
            senha=generate_password_hash('SUPORTE26@'),
            perfil='MASTER',
            acesso_dashboard=True,
            acesso_configuracoes=True,
            acesso_relatorios=True
        )
        db.session.add(master)

        # Criar usuario admin
        admin = Usuario(
            nome='admin',
            senha=generate_password_hash('123'),
            perfil='ADMIN',
            acesso_dashboard=True,
            acesso_configuracoes=True,
            acesso_relatorios=True
        )
        db.session.add(admin)

        # Criar formas de pagamento padrão
        formas = [
            'Dinheiro', 'Crédito', 'Débito', 'PIX', 'Cartão (Voucher)',
            'Conta Assinada', 'PG Online', 'Link de Pagamento',
            'Transferência', 'Depósito', 'Boleto', 'Cheque',
            'Vale Refeição', 'Vale Alimentação', 'Cortesia'
        ]
        for forma in formas:
            db.session.add(FormaPagamento(nome=forma))
        
        # Criar bandeiras padrão
        bandeiras = [
            'Visa', 'Mastercard', 'Elo', 'American Express',
            'Hipercard', 'Diners Club', 'Discover', 'Aura',
            'Cabal', 'Banescard', 'Good Card', 'Sodexo',
            'Ticket', 'VR', 'Alelo'
        ]
        for bandeira in bandeiras:
            db.session.add(BandeiraCartao(nome=bandeira))
        
        # Criar categorias padrão
        categorias = [
            ('Aluguel', 'FIXA'),
            ('Condomínio', 'FIXA'),
            ('Água', 'FIXA'),
            ('Luz', 'FIXA'),
            ('Internet', 'FIXA'),
            ('Telefonia', 'FIXA'),
            ('Contabilidade', 'FIXA'),
            ('Sistema/Software', 'FIXA'),
            ('Produtos', 'VARIAVEL'),
            ('Embalagens', 'VARIAVEL'),
            ('Gás', 'VARIAVEL'),
            ('Manutenção', 'VARIAVEL'),
            ('Limpeza', 'VARIAVEL'),
            ('Marketing', 'VARIAVEL'),
            ('Fretado/Entrega', 'VARIAVEL'),
            ('Comissão Motoboy', 'VARIAVEL'),
            ('Taxas Cartão', 'SAIDA'),
            ('Taxas Plataforma', 'SAIDA'),
            ('Impostos', 'SAIDA'),
            ('Passagem', 'SAIDA'),
            ('Multas', 'SAIDA'),
            ('Outros', 'SAIDA')
        ]
        for nome, tipo in categorias:
            db.session.add(CategoriaDespesa(nome=nome, tipo=tipo))
        
        # Criar motoboys padrão
        motoboys = ['João', 'Maria', 'Pedro']
        for nome in motoboys:
            db.session.add(Motoboy(nome=nome, taxa_padrao=5.00))
        
        db.session.commit()
        print("✅ Banco de dados inicializado com sucesso!")

@app.route('/criar-chave-teste')
def criar_chave_teste():
    """Criar uma chave de teste"""
    import random
    import string
    
    # Gerar chave no formato XXXX-XXXX-XXXX-XXXX
    chave = '-'.join(
        ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
        for _ in range(4)
    )
    
    # Criar licença de teste
    licenca = Licenca(
        email='teste@email.com',
        chave_ativacao=chave,
        data_expiracao=datetime.utcnow() + timedelta(days=365),
        status='ATIVA'
    )
    db.session.add(licenca)
    db.session.commit()
    
    return f"""
    <h1>Chave de Teste Criada!</h1>
    <h3>E-mail: teste@email.com</h3>
    <h3>Chave: <code>{chave}</code></h3>
    <button onclick="navigator.clipboard.writeText('{chave}')">Copiar Chave</button>
    <br><br>
    <a href="/ativacao">Ir para tela de ativação</a>
    """

@app.route('/ver-chaves')
def ver_chaves():
    """Rota temporária para ver chaves de licença"""
    licencas = Licenca.query.all()
    
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Chaves de Licença</title>
        <style>
            body { font-family: Arial; padding: 20px; }
            .chave { background: #f0f0f0; padding: 10px; margin: 10px; border-radius: 5px; }
            code { font-size: 16px; color: #d63384; }
        </style>
    </head>
    <body>
        <h1>Chaves de Licença Cadastradas</h1>
    """
    
    if licencas:
        for licenca in licencas:
            html += f"""
            <div class="chave">
                <strong>ID:</strong> {licenca.id}<br>
                <strong>E-mail:</strong> {licenca.email}<br>
                <strong>Chave:</strong> <code>{licenca.chave_ativacao}</code><br>
                <strong>Status:</strong> {licenca.status}<br>
                <strong>Ativa:</strong> {licenca.ativo}<br>
                <button onclick="navigator.clipboard.writeText('{licenca.chave_ativacao}')">Copiar Chave</button>
            </div>
            <hr>
            """
    else:
        html += "<p>Nenhuma licença cadastrada!</p>"
    
    html += """
        <script>
            function copiarTodas() {
                let chaves = [];
                document.querySelectorAll('code').forEach(code => {
                    chaves.push(code.textContent);
                });
                navigator.clipboard.writeText(chaves.join('\\n'));
                alert('Todas as chaves copiadas!');
            }
        </script>
        <button onclick="copiarTodas()">Copiar Todas as Chaves</button>
    </body>
    </html>
    """
    
    return html



# ==================== RUN ====================

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Correção automática: Adicionar coluna acesso_relatorios se não existir
        try:
            from sqlalchemy import text
            with db.engine.connect() as conn:
                conn.execute(text('ALTER TABLE usuario ADD COLUMN acesso_relatorios BOOLEAN DEFAULT 0'))
                conn.commit()
            print("✅ Coluna 'acesso_relatorios' adicionada automaticamente!")
        except Exception:
            pass  # Coluna já existe ou erro ao adicionar

        # Criar admin se não existir
        if not Usuario.query.filter_by(perfil='ADMIN').first():
            from werkzeug.security import generate_password_hash
            admin = Usuario(
                nome='admin',
                senha=generate_password_hash('admin123'),
                perfil='ADMIN',
                ativo=True,
                acesso_configuracoes=True
            )
            db.session.add(admin)
            db.session.commit()
            print('✅ Admin padrão criado')
            
        print("Banco atualizado!")
    app.run(
        host='0.0.0.0',
        port=int(os.environ.get('PORT', 5000)),
        debug=bool(os.environ.get('FLASK_DEBUG', ''))
    )
